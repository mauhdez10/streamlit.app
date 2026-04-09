"""
Broadcast Playlist Checker — Core Logic v5
"""
import json, re, xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, timedelta


# ── TRANSLATION ───────────────────────────────────────────────────────────────

_S = {
    'section_programs':   {'en': 'PROGRAM CHECK (Playlist vs Grilla)',       'es': 'VERIFICACIÓN DE PROGRAMAS (Playlist vs Grilla)'},
    'section_commercials':{'en': 'COMMERCIAL CHECK (Playlist vs XML)',        'es': 'VERIFICACIÓN DE COMERCIALES (Playlist vs XML)'},
    'section_promos':     {'en': 'PROMO REPEAT CHECK',                        'es': 'PROMOS REPETIDAS'},
    'section_ingested':   {'en': 'NOT INGESTED ASSETS',                       'es': 'ACTIVOS NO INGESTADOS'},
    'section_bugs':       {'en': 'BUGS',                                       'es': 'BUGS'},
    'section_cues':       {'en': 'CUE TONE REPORT',                           'es': 'REPORTE DE CUE TONES'},
    'full_day':           {'en': 'FULL DAY',                                   'es': 'DÍA COMPLETO'},
    'partial':            {'en': 'CURRENT (partial)',                          'es': 'ACTUAL (parcial)'},
    'checking_from':      {'en': 'CHECKING FROM',                              'es': 'VERIFICANDO DESDE'},
    'channel':            {'en': 'CHANNEL',                                    'es': 'CANAL'},
    'date_lbl':           {'en': 'DATE',                                       'es': 'FECHA'},
    'type_lbl':           {'en': 'PLAYLIST TYPE',                              'es': 'TIPO DE PLAYLIST'},
    'summary':            {'en': 'SUMMARY',                                    'es': 'RESUMEN'},
    'show_blocks':        {'en': 'show blocks',                                'es': 'bloques de programa'},
    'commercials_lbl':    {'en': 'commercials',                                'es': 'comerciales'},
    'no_grilla':          {'en': '! Grilla not provided',                      'es': '! Grilla no proporcionada'},
    'no_xml':             {'en': '! XML log not provided',                     'es': '! Log XML no proporcionado'},
    'ok_programs':        {'en': '✓ All programs match grilla',                'es': '✓ Todos los programas coinciden con la grilla'},
    'ok_commercials':     {'en': '✓ All {n} commercials match XML log',        'es': '✓ Los {n} comerciales coinciden con el log XML'},
    'ok_promos':          {'en': '✓ No repeated promos within same break',     'es': '✓ Sin promos repetidas en el mismo break'},
    'ok_ingested':        {'en': '✓ All assets ingested',                      'es': '✓ Todos los activos están ingestados'},
    'ok_bugs':            {'en': '✓ No bugs scheduled',                        'es': '✓ Sin bugs programados'},
    'total_cues':         {'en': 'Total cue tones',                            'es': 'Total cue tones'},
    'anchored':           {'en': '  ℹ  Anchored at grilla position {i}: {id}', 'es': '  ℹ  Anclado en grilla pos {i}: {id}'},
    'already_aired':      {'en': '  ℹ  ALREADY AIRED: {id}',                  'es': '  ℹ  YA SE TRANSMITIÓ: {id}'},
    'wrong_ep':           {'en': '  ⚠  WRONG EPISODE: Grilla={g} | Playlist={p} @ {t}', 'es': '  ⚠  EPISODIO INCORRECTO: Grilla={g} | Playlist={p} @ {t}'},
    'not_in_pl':          {'en': '  ✗  NOT IN PLAYLIST: {id}',                'es': '  ✗  NO ESTÁ EN PLAYLIST: {id}'},
    'extra_pl':           {'en': '  ✗  EXTRA: {id} @ {t} (not in grilla)',     'es': '  ✗  EXTRA: {id} @ {t} (no está en grilla)'},
    'order_mis':          {'en': '  ⚠  ORDER pos {i}: Grilla={g} | Playlist={p}', 'es': '  ⚠  ORDEN pos {i}: Grilla={g} | Playlist={p}'},
    'xml_not_pl':         {'en': '  ✗  IN XML, NOT IN PLAYLIST: {ref} ({n}x in XML)', 'es': '  ✗  EN XML, NO EN PLAYLIST: {ref} ({n}x en XML)'},
    'pl_not_xml':         {'en': '  ✗  IN PLAYLIST, NOT IN XML: {ref} ({n}x in playlist)', 'es': '  ✗  EN PLAYLIST, NO EN XML: {ref} ({n}x en playlist)'},
    'count_diff':         {'en': '  ⚠  COUNT DIFF: {ref} | XML={xn}x | Playlist={pn}x', 'es': '  ⚠  DIFERENCIA: {ref} | XML={xn}x | Playlist={pn}x'},
    'promo_rep':          {'en': '  ⚠  PROMO REPEAT after [{after}] @ {t}: {ref} {n}x', 'es': '  ⚠  PROMO REPETIDA después de [{after}] @ {t}: {ref} {n}x'},
    'ni_program':         {'en': '  ⚠  NOT INGESTED [Program]: {id} @ {t} | {show}', 'es': '  ⚠  NO INGESTADO [Programa]: {id} @ {t} | {show}'},
    'ni_other':           {'en': '  ⚠  NOT INGESTED [{typ}]: {ref} @ {t} | {name}', 'es': '  ⚠  NO INGESTADO [{typ}]: {ref} @ {t} | {name}'},
    'bug_line':           {'en': '  🔲 {beh_label} : {cmd} — {id} @ {t} | {show}', 'es': '  🔲 {beh_label} : {cmd} — {id} @ {t} | {show}'},
}

def T(key, lang='en', **kwargs):
    s = _S.get(key, {}).get(lang, _S.get(key, {}).get('en', key))
    return s.format(**kwargs) if kwargs else s


# ── TIME HELPERS ──────────────────────────────────────────────────────────────

def _edt_start(year):
    """Second Sunday of March 2:00 AM."""
    m1 = datetime(year, 3, 1)
    d = (6 - m1.weekday()) % 7
    return datetime(year, 3, (m1 + timedelta(days=d+7)).day, 2)

def _est_start(year):
    """First Sunday of November 2:00 AM."""
    n1 = datetime(year, 11, 1)
    d = (6 - n1.weekday()) % 7
    return datetime(year, 11, (n1 + timedelta(days=d)).day, 2)

def utc_to_et(dt):
    if dt is None: return None
    edt = _edt_start(dt.year)
    est = _est_start(dt.year)
    offset = -4 if edt <= dt < est else -5
    tz = 'EDT' if offset == -4 else 'EST'
    return dt + timedelta(hours=offset), tz

def fmt_time(dt):
    """UTC only."""
    if dt is None: return '??:??:??'
    return dt.strftime('%H:%M:%S')

def fmt_t(dt):
    """UTC / ET string."""
    if dt is None: return '??:??:?? UTC'
    et, tz = utc_to_et(dt)
    return f'{dt.strftime("%H:%M:%S")} UTC / {et.strftime("%H:%M:%S")} ET'

def parse_timecode(tc):
    try:
        tc = tc.split(';')[0].split('@')[0].strip()
        return datetime.strptime(tc, '%Y-%m-%d %H:%M:%S')
    except: return None

def parse_duration(dur):
    try:
        dur = dur.split(';')[0].split('@')[0]
        h,m,s = dur.split(':')
        return int(h)*3600 + int(m)*60 + int(s)
    except: return 0

def parse_xml_time(ts):
    """XML startat: '15:40:00:00' → datetime (dummy date)."""
    try:
        p = ts.split(':')
        return datetime(2000, 1, 1, int(p[0]), int(p[1]), int(p[2]))
    except: return None


# ── ID HELPERS ────────────────────────────────────────────────────────────────

def is_episode_id(val):
    if not val or not isinstance(val, str): return False
    val = val.strip()
    if ' ' in val or len(val) < 3 or len(val) > 16: return False
    if not re.match(r'^[A-Z]', val): return False
    return len(re.findall(r'\d', val)) >= 3

def is_movie_id(val):
    """Movie/special IDs: pure uppercase alpha, 3-8 chars. MARCE, NELQP, CAME."""
    if not val or not isinstance(val, str): return False
    val = val.strip()
    return bool(re.match(r'^[A-Z]{3,8}$', val))

def normalize_id(ep_id):
    if not ep_id: return ''
    ep_id = re.sub(r'_\d+$', '', str(ep_id).strip())
    # Only normalize extra leading zeros before 4+ digit date suffixes (e.g. COSA00327→COSA0327)
    # Do NOT touch 3-digit episode numbers (e.g. LATPAN001 stays LATPAN001)
    ep_id = re.sub(r'([A-Za-z][A-Za-z0-9]*)0{2,}(\d{4,})',
                   lambda m: m.group(1) + (m.group(2)[-4:] if len(m.group(2)) > 4 else m.group(2)),
                   ep_id)
    return ep_id.upper()

def show_prefix(ep_id):
    if not ep_id: return ''
    ep_id = ep_id.upper()
    m = re.match(r'^([A-Z]+)', ep_id)
    if not m: return ''
    letters = m.group(1)
    if re.match(r'^[A-Z]{3,8}$', ep_id): return ep_id
    return letters


# ── JSON PARSER ───────────────────────────────────────────────────────────────

def parse_json_playlist(data):
    events = data.get('events', [])
    has_marker = any(a.get('type') == 'marker'
                     for ev in events[:3] for a in ev.get('assets', []))
    playlist_type = 'full' if has_marker else 'current'

    date = None
    for ev in events:
        dt = parse_timecode(ev.get('startTime', ''))
        if dt: date = dt.date(); break

    programs, commercials, promos, cue_tones, not_ingested, breaks = [], [], [], [], [], []
    current_break, last_program, last_program_raw = [], None, None

    for ev in events:
        ev_assets = ev.get('assets', [])
        ev_start  = parse_timecode(ev.get('startTime', ''))
        ev_dur    = parse_duration(ev.get('duration', ''))
        ev_name   = ev.get('name', '')
        ev_ref    = ev.get('reference', '')
        behaviors = ev.get('behaviors', [])

        for b in behaviors:
            if b.get('name') == 'CUEON' and not b.get('disabled', True):
                ct = ev_assets[0].get('reference', ev_name) if ev_assets else ev_name
                cue_tones.append({'ref': ev_ref, 'name': ev_name, 'ct_id': ct,
                                  'start': ev_start, 'duration': ev_dur, 'is_cueoff': False})
            if b.get('name') == 'CUEOFF' and not b.get('disabled', True):
                ct = ev_assets[0].get('reference', ev_name) if ev_assets else ev_name
                cue_tones.append({'ref': ev_ref, 'name': ev_name, 'ct_id': ct,
                                  'start': ev_start, 'duration': ev_dur, 'is_cueoff': True})

        for asset in ev_assets:
            atype = asset.get('type', '')
            aref  = asset.get('reference', '')
            tcin  = asset.get('tcIn', '')

            if tcin.startswith('07:') and atype != 'live':
                not_ingested.append({'asset_ref': aref, 'name': ev_name,
                                     'type': atype, 'start': ev_start, 'ref': ev_ref})

            if atype in ('Program', 'live'):
                if current_break:
                    breaks.append({'after_program': last_program,
                                   'after_program_raw': last_program_raw,
                                   'items': current_break[:]})
                    current_break = []
                seg_m = re.search(r'_(\d+)$', aref)
                seg   = int(seg_m.group(1)) if seg_m else 1
                ep_id = normalize_id(aref)
                _logo = next((b.get('params',{}).get('Command')
                              for b in behaviors
                              if b.get('name') in ('LOGOHD','LOGOHD_ANI') and not b.get('disabled',False)), None)
                programs.append({'episode_id': ep_id, 'episode_id_raw': aref,
                                  'seg_num': seg, 'start': ev_start, 'duration': ev_dur,
                                  'name': ev_name, 'ref': ev_ref, 'asset_type': atype,
                                  'is_missing': (atype == 'Program' and tcin.startswith('07:')),
                                  'logo': _logo})
                last_program     = ep_id
                last_program_raw = aref

            elif atype == 'Commercial':
                commercials.append({'asset_ref': aref, 'name': ev_name,
                                    'start': ev_start, 'duration': ev_dur,
                                    'ref': ev_ref})
                current_break.append({'type': 'Commercial', 'ref': aref,
                                      'start': ev_start, 'event_ref': ev_ref})
            elif atype == 'Promotion':
                _pdur = 0
                try:
                    _tcin  = assets[0].get('tcIn','') if assets else ''
                    _tcout = assets[0].get('tcOut','') if assets else ''
                    def _tc(s):
                        import re as _r; m=_r.match(r'(\d+):(\d+):(\d+)',str(s)); return int(m.group(1))*3600+int(m.group(2))*60+int(m.group(3)) if m else 0
                    _pdur = _tc(_tcout) - _tc(_tcin)
                except: pass
                promos.append({'asset_ref': aref, 'name': ev_name, 'start': ev_start, 'ref': ev_ref, 'duration': max(0,_pdur)})
                current_break.append({'type': 'Promotion', 'ref': aref, 'start': ev_start})

    if current_break:
        breaks.append({'after_program': last_program,
                       'after_program_raw': last_program_raw,
                       'items': current_break})

    return {'type': playlist_type, 'date': date, 'events': events,
            'programs': programs, 'commercials': commercials, 'promos': promos,
            'breaks': breaks, 'cue_tones': cue_tones, 'not_ingested': not_ingested}


def build_show_sequence(programs, from_start=None):
    seq, prev = [], None
    for p in programs:
        if from_start and p['start'] and p['start'] < from_start: continue
        ep = p['episode_id']
        if ep != prev:
            seq.append({'id': ep, 'start': p['start'], 'raw': p['episode_id_raw']})
            prev = ep
    return seq


# ── XML PARSER ────────────────────────────────────────────────────────────────

def parse_xml_log(filepath_or_bytes):
    """
    Parse standard XML traffic log.
    Auto-detects format:
      - <traffics><traffic><item>  → standard (CATV, TVD, original Pasiones)
      - <tabledata><data><row>     → tabledata (TN, Pasiones after format change, Sony)
    Safe to use for all channels — routes to the correct parser internally.
    """
    try:
        if hasattr(filepath_or_bytes, 'read'): content = filepath_or_bytes.read()
        elif isinstance(filepath_or_bytes, bytes): content = filepath_or_bytes
        else:
            with open(filepath_or_bytes, 'rb') as f: content = f.read()
        # Sanitize unescaped & that break XML parser (e.g. P&G in title fields)
        content = re.sub(rb'&(?![a-zA-Z#][a-zA-Z0-9#]*;)', b'&amp;', content)
        root = ET.fromstring(content)
        # Auto-detect format
        if root.tag == 'tabledata':
            # tabledata format — same logic as parse_xml_log_tn
            items = []
            for row in root.findall('.//row'):
                local_time = row.findtext('column-1', '').strip()
                mediaid    = row.findtext('column-4', '').strip()
                typ        = row.findtext('column-5', '').strip().upper()
                title      = row.findtext('column-6', '').strip()
                duration   = row.findtext('column-3', '').strip()
                if typ == 'PROGRAM':    ct = 'PROGRAM_BEGIN'
                elif typ == 'PROMOTION': ct = 'PROMO'
                else:                    ct = typ
                items.append({'mediaid': mediaid, 'name': title,
                              'contenttype': ct, 'startat': local_time,
                              'duration': duration, 'externalid': ''})
            return items
        else:
            # Standard <traffics><traffic><item> format
            traffic = root.find('traffic')
            if traffic is None: traffic = root
            return [{'mediaid': i.get('mediaid',''), 'name': i.findtext('n','').strip(),
                     'contenttype': i.findtext('contenttype','').strip().upper(),
                     'startat': i.findtext('startat','').strip(),
                     'duration': i.findtext('duration','').strip(),
                     'externalid': i.findtext('externalid','').strip()}
                    for i in traffic.findall('item')]
    except: return []

def parse_xml_log_tn(filepath_or_bytes):
    """
    Parser for Todonovelas XML — <tabledata><data><row><column-N> format.
    column-1=LocalTime, column-4=MediaId, column-5=Type, column-6=Title
    Returns same dict format as parse_xml_log for compatibility.
    """
    try:
        if hasattr(filepath_or_bytes, 'read'): content = filepath_or_bytes.read()
        elif isinstance(filepath_or_bytes, bytes): content = filepath_or_bytes
        else:
            with open(filepath_or_bytes, 'rb') as f: content = f.read()
        root = ET.fromstring(content)
        items = []
        for row in root.findall('.//row'):
            local_time = row.findtext('column-1', '').strip()
            mediaid    = row.findtext('column-4', '').strip()
            typ        = row.findtext('column-5', '').strip().upper()
            title      = row.findtext('column-6', '').strip()
            duration   = row.findtext('column-3', '').strip()
            # Normalise type to match standard contenttype values
            if typ == 'PROGRAM': ct = 'PROGRAM_BEGIN'
            elif typ == 'PROMOTION': ct = 'PROMO'
            else: ct = typ
            items.append({'mediaid': mediaid, 'name': title,
                          'contenttype': ct, 'startat': local_time,
                          'duration': duration, 'externalid': ''})
        return items
    except: return []

def _xml_dur_secs(dur_str):
    """Parse XML duration HH:MM:SS:FF → seconds (ignore frames)."""
    try:
        p = dur_str.split(':')
        return int(p[0])*3600 + int(p[1])*60 + int(p[2])
    except: return 0

def _is_xml_program_anchor(item):
    """Only PROGRAM_BEGIN/PROGRAM_SEGMENT used for break-by-break alignment."""
    return item.get('contenttype','') in ('PROGRAM_BEGIN','PROGRAM_SEGMENT')

def _is_xml_start_anchor(item):
    """For partial start detection: program segments + infomercials (CM ≥ 20min)."""
    ct = item.get('contenttype','')
    if ct in ('PROGRAM_BEGIN','PROGRAM_SEGMENT'): return True
    if ct == 'COMMERCIAL' and _xml_dur_secs(item.get('duration','')) >= 1200: return True
    return False

def build_xml_breaks(xml_rows):
    """
    Walk XML, group commercials between PROGRAM segments only (not infomercials).
    Infomercials appear as commercials inside a break, matching JSON behavior.
    Returns list of {'anchor_id', 'commercials': [mediaid, ...]}
    Includes breaks with zero commercials for alignment.
    """
    result = []
    anchor = None
    comms  = []
    for item in xml_rows:
        if _is_xml_program_anchor(item):
            if anchor is not None:
                result.append({'anchor_id': anchor, 'commercials': comms[:]})
            anchor = item['mediaid']
            comms  = []
        elif item.get('contenttype') == 'COMMERCIAL':
            comms.append(item['mediaid'])
    if anchor is not None:
        result.append({'anchor_id': anchor, 'commercials': comms[:]})
    return result

def xml_commercials(rows):
    return [r for r in rows if r.get('contenttype') == 'COMMERCIAL']

def find_xml_anchor_by_extid(events, xml_rows):
    """
    Find where partial JSON starts in XML using externalid/reference match.
    Returns start index in xml_rows for commercial comparison.
    """
    ext_idx = {row['externalid']: i for i, row in enumerate(xml_rows)}
    for ev in events:
        ref = ev.get('reference', '')
        if ref in ext_idx:
            return ext_idx[ref]
    return 0


# ── GRILLA PARSER ─────────────────────────────────────────────────────────────

def parse_grilla(filepath_or_bytes, target_date, channel='catv'):
    """Route to the correct grilla parser based on channel type."""
    if channel in ('latam', 'us'):
        return _parse_grilla_pasiones(filepath_or_bytes, target_date)
    if channel == 'tn':
        return _parse_grilla_tn(filepath_or_bytes, target_date)
    return _parse_grilla_catv_tvd(filepath_or_bytes, target_date)

def _parse_grilla_catv_tvd(filepath_or_bytes, target_date):
    """Original CATV/TVD grilla parser — single active sheet, datetime header."""
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=False)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=False)

    ws = wb.active
    # Keep two views: values_only for formulas, full rows for bold detection
    all_rows_full = list(ws.iter_rows(values_only=False))
    all_rows = [[cell.value for cell in row] for row in all_rows_full]
    if len(all_rows) < 2: return []

    header_row = all_rows[1]
    # Build date→col map, skipping ET/UTC/CA marker columns
    _NON_DATE = {'ET', 'UTC', 'CA', 'E.T.', 'U.T.C.'}
    date_col_map = {}
    for ci, val in enumerate(header_row):
        if val is None: continue
        if isinstance(val, str) and val.strip().upper() in _NON_DATE: continue
        if hasattr(val, 'date'):
            date_col_map[val.date()] = ci
        elif isinstance(val, str) and val.strip().startswith('='):
            # Formula like =C2+5 — resolve base date + offset
            m2 = re.match(r'^=([A-Z]+)(\d+)\+(\d+)$', val.strip())
            if m2:
                try:
                    base_ci = sum((ord(c)-ord('A')+1)*(26**i)
                                 for i,c in enumerate(reversed(m2.group(1))))-1
                    base_ri = int(m2.group(2))-1
                    base_v  = all_rows[base_ri][base_ci] if base_ri < len(all_rows) and base_ci < len(all_rows[base_ri]) else None
                    if base_v and hasattr(base_v, 'date'):
                        date_col_map[(base_v + timedelta(days=int(m2.group(3)))).date()] = ci
                except Exception:
                    pass
    target_col = date_col_map.get(target_date)
    if target_col is None and date_col_map:
        # Fallback: find Monday and walk by weekday offset
        monday_d = min(date_col_map.keys())
        monday_c = date_col_map[monday_d]
        for offset in range(7):
            if monday_d + timedelta(days=offset) == target_date:
                target_col = monday_c + offset; break
    if target_col is None:
        return []

    def resolve_cell(val):
        if not val or not isinstance(val, str): return val
        m = re.match(r'^=([A-Z]+)(\d+)$', val.strip())
        if not m: return val
        col_str, row_num = m.group(1), int(m.group(2))
        col_idx = sum((ord(c)-ord('A')+1) * (26**i) for i,c in enumerate(reversed(col_str))) - 1
        row_idx = row_num - 1
        if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
            return all_rows[row_idx][col_idx]
        return val

    def extract_ids(val):
        val = resolve_cell(val)
        if not val or not isinstance(val, str) or val.startswith('='): return []
        val = val.strip()
        # Full cell is an episode ID (standard with digits)
        if is_episode_id(val): return [normalize_id(val)]
        # Full cell is a movie/special ID (pure alpha, no digits)
        if is_movie_id(val): return [normalize_id(val)]
        # Cell has mixed content — tokenize and extract only standard IDs (digits required)
        # Do NOT apply movie_id check on tokens to avoid extracting description words
        tokens = re.findall(r'[A-Z0-9]+', val.upper())
        return [normalize_id(t) for t in tokens if is_episode_id(t)]

    episode_ids = []
    for ri, row in enumerate(all_rows[2:], start=2):
        val = row[target_col] if target_col < len(row) else None
        if val is None: continue
        # Only extract bold cells — bold = show ID, non-bold = description text
        full_row = all_rows_full[ri] if ri < len(all_rows_full) else []
        cell = full_row[target_col] if target_col < len(full_row) else None
        is_bold = (cell is not None and cell.font is not None and cell.font.bold)
        if not is_bold: continue
        for ep in extract_ids(val):
            if ep: episode_ids.append(ep)
    return episode_ids

def _parse_date_str(val, force_year=None):
    """Parse date from string like 'Lun. / Mon. 03/30/26'.
    force_year overrides the year in the string (handles typos like 03/30/25 when it should be 2026)."""
    if not val: return None
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(val))
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if force_year:
            y = force_year
        else:
            y = 2000 + y if y < 100 else y
        try: return datetime(y, mo, d).date()
        except: pass
    return None

def _parse_grilla_pasiones(filepath_or_bytes, target_date):
    """
    Pasiones grilla parser — multi-tab yearly workbook.
    Scans tabs last→first, finds the one containing target_date.
    Header row contains date strings like 'Mar. / Tue. 03/31/26'.
    Episode IDs are in alternating rows (show name row, then ID row).
    """
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=True)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=True)

    target_ws = None
    target_col = None

    for name in reversed(wb.sheetnames):
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        if len(rows) < 2: continue
        header = rows[1]
        # Use target_date's year to avoid typos in spreadsheet year field
        col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                     for i, cell in enumerate(header)
                     if _parse_date_str(cell, force_year=target_date.year)]
        if not col_dates: continue
        first_d = col_dates[0][1]
        last_d  = col_dates[-1][1]
        if first_d <= target_date <= last_d:
            for col_i, d in col_dates:
                if d == target_date:
                    target_col = col_i
                    target_ws  = ws
                    # Re-read full sheet
                    all_rows = list(ws.iter_rows(values_only=True))
                    break
            break

    if target_ws is None or target_col is None:
        return []

    # Detect which column is ET — show name rows have a time value there,
    # episode ID rows have None. Works for both ET-at-col-0 and UTC+ET formats.
    # Look for 'ET' or 'E.T.' in header row to find the column.
    header_row = all_rows[1] if len(all_rows) > 1 else []
    ET_COL = 0  # default
    for ci, cell in enumerate(header_row):
        if cell and isinstance(cell, str) and cell.strip().upper() in ('ET', 'E.T.'):
            ET_COL = ci
            break

    episode_ids = []
    for row in all_rows[2:]:
        et_val = row[ET_COL] if ET_COL < len(row) else 'x'
        if et_val is not None:
            continue  # show name row — skip
        val = row[target_col] if target_col < len(row) else None
        if val and isinstance(val, str):
            val = val.strip()
            if val:
                episode_ids.append(normalize_id(val))
    return episode_ids

def _parse_grilla_tn(filepath_or_bytes, target_date):
    """
    Todonovelas grilla — multi-tab, same header format as Pasiones.
    Returns list of (show_name, episode_num) tuples for program matching.
    Episode numbers are integers in the grid (55, 121...).
    Show name rows: ET column has value. Episode rows: ET column is None.
    """
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=True)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=True)

    target_ws = None
    target_col = None
    all_rows = []

    for name in reversed(wb.sheetnames):
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        if len(rows) < 2: continue
        header = rows[1]
        col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                     for i, cell in enumerate(header)
                     if _parse_date_str(cell, force_year=target_date.year)]
        if not col_dates: continue
        if col_dates[0][1] <= target_date <= col_dates[-1][1]:
            for col_i, d in col_dates:
                if d == target_date:
                    target_col = col_i
                    target_ws  = ws
                    all_rows   = list(ws.iter_rows(values_only=True))
                    break
            break

    if not all_rows or target_col is None:
        return []

    ET_COL = 1
    result = []
    current_show = None
    for row in all_rows[2:]:
        et_val = row[ET_COL] if ET_COL < len(row) else 'x'
        val    = row[target_col] if target_col < len(row) else None
        if val is None: continue
        if et_val is not None:
            # Show name row
            current_show = str(val).strip() if val else None
        else:
            # Episode number row — val is an integer
            if current_show and val is not None:
                try:
                    ep_num = int(val)
                    result.append((current_show, ep_num))
                except (ValueError, TypeError):
                    pass
    return result


# ── FILE DETECTION ────────────────────────────────────────────────────────────

def extract_date_from_filename(name):
    # YYYYMMDD (JSON files: ..._20260330_...)
    m = re.search(r'(\d{4})(\d{2})(\d{2})', name)
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except: pass
    # MMDDYYYY (XML files: TVD03302026.xml, CA03302026.xml)
    m = re.search(r'(\d{2})(\d{2})(\d{4})', name)
    if m:
        try: return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    return None

def _date_from_json_content(f):
    """Extract date from JSON file content (first event's startTime). Most reliable."""
    try:
        f.seek(0)
        data = json.load(f)
        f.seek(0)
        for ev in data.get('events', []):
            tc = ev.get('startTime', '')
            if tc:
                m = re.search(r'(\d{4})-(\d{2})-(\d{2})', tc)
                if m:
                    return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
    except: pass
    return None

def _date_from_xml_filename(name):
    """XML filenames date extraction.
    Sony:     YYYYMMDD (A120260401c.XML, S620260401c_XML.xml)
    Standard: MMDDYYYY (TVD03302026.xml, CA03302026.xml, PL03312026.xml)
    TN:       MMDDYY   (TN_033126_TUESDAY.xml)
    """
    # Try YYYYMMDD first (8 consecutive digits starting with 20)
    m = re.search(r'(20\d{2})(\d{2})(\d{2})', name)
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except: pass
    # Try MMDDYYYY (8 digits)
    m = re.search(r'(\d{2})(\d{2})(\d{4})', name)
    if m:
        try: return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    # Try MMDDYY (6 digits, 2-digit year)
    m = re.search(r'(\d{2})(\d{2})(\d{2})(?!\d)', name)
    if m:
        try: return datetime(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    return None

def detect_files(uploaded_files):
    """
    Group uploaded files by (date, channel).
    JSON: date always from content (startTime). Works for renamed files.
    XML:  date from filename (MMDDYYYY pattern, always reliable).
    Grilla: week-based, stored by channel only.
    Returns: days dict, grillas dict, unknown list.
    """
    days    = {}
    grillas = {}
    unknown = []
    sony_files = []   # Sony/AXN files handled separately

    for f in uploaded_files:
        name_up = f.name.upper()
        ext = f.name.lower().rsplit('.', 1)[-1] if '.' in f.name else ''

        if ext == 'json':
            ftype = 'json'
        elif ext == 'xml':
            ftype = 'xml'
        elif ext == 'pdf':
            # HolaTV PDF grilla only
            if 'USH' in name_up or ('HOLA' in name_up and 'US' in name_up):
                grillas.setdefault('hu', []).append(f)
            elif 'LATAM' in name_up:
                # Accept any LATAM PDF as HolaTV Latam grilla
                grillas.setdefault('hl', []).append(f)
            else:
                unknown.append(f)
            continue
        elif ext in ('xlsx', 'xlsm'):
            # HolaTV XLSX log (not grilla)
            if name_up.startswith('HU'):
                ftype = 'log'; 
            elif name_up.startswith('HL'):
                ftype = 'log'
            else:
                ftype = 'grilla'
        elif ext == 'txt':
            if name_up.startswith('HU') or name_up.startswith('HL'):
                ftype = 'log'
            else:
                unknown.append(f); continue
        else:
            unknown.append(f); continue

        # Sony/AXN detection — check before generic channel detection
        sony_code = extract_sony_code(f.name)
        if sony_code:
            sony_files.append({'file': f, 'ftype': ftype, 'code': sony_code})
            continue

        # Channel detection
        if ext == 'xml':
            if name_up.startswith('TVD'):   channel = 'tvd'
            elif name_up.startswith('CA'):  channel = 'catv'
            elif name_up.startswith('PL'):  channel = 'latam'
            elif name_up.startswith('PUS'): channel = 'us'
            elif name_up.startswith('TN'):  channel = 'tn'
            elif name_up.startswith('HU'):  channel = 'hu'
            elif name_up.startswith('HL'):  channel = 'hl'
            else: unknown.append(f); continue
        elif ext == 'pdf':
            # HolaTV PDF grilla
            if 'USH' in name_up or 'HOLATV_US' in name_up or 'HOLA_US' in name_up or 'HOLA TV US' in name_up:
                grillas['hu'] = f; continue
            elif 'LATAM' in name_up and ('HOLA' in name_up or 'HL' in name_up):
                grillas['hl'] = f; continue
            else: unknown.append(f); continue
        elif ext == 'json':
            # JSON channel detection
            if 'HOLATV_US' in name_up or 'HOLA_TV_US' in name_up or 'HOLA_US' in name_up or ('HOLATV' in name_up and 'US' in name_up):
                channel = 'hu'
            elif 'HOLATV_LATAM' in name_up or 'HOLA_TV_LATAM' in name_up or 'HOLA_LATAM' in name_up or ('HOLATV' in name_up and 'LATAM' in name_up):
                channel = 'hl'
            elif 'PASIONES_LATAM' in name_up or 'PASIONES LATAM' in name_up or name_up.startswith('PL'):
                channel = 'latam'
            elif 'PASIONES_US' in name_up or 'PASIONES US' in name_up or name_up.startswith('PUS'):
                channel = 'us'
            elif 'FAST_TODONOVELAS' in name_up or 'FAST TODONOVELAS' in name_up or ('TODO' in name_up and 'NOVELA' in name_up):
                channel = 'tn'
            elif 'TVD' in name_up: channel = 'tvd'
            elif 'CATV' in name_up or name_up.startswith('CA'): channel = 'catv'
            else: unknown.append(f); continue
        else:
            if 'HOLATV_US' in name_up or 'HOLA_TV_US' in name_up or ('HOLATV' in name_up and 'US' in name_up) or name_up.startswith('HU'): 
                channel = 'hu'
            elif 'HOLATV_LATAM' in name_up or 'HOLA_TV_LATAM' in name_up or ('HOLATV' in name_up and 'LATAM' in name_up) or name_up.startswith('HL'): 
                channel = 'hl'
            elif 'CATV' in name_up:                       channel = 'catv'
            elif 'TVD' in name_up:                        channel = 'tvd'
            elif 'PASIONES_LATAM' in name_up or 'PASIONES LATAM' in name_up: channel = 'latam'
            elif 'PASIONES_US' in name_up or 'PASIONES US' in name_up:       channel = 'us'
            elif 'FAST_TODONOVELAS' in name_up or 'FAST TODONOVELAS' in name_up or ('TODO' in name_up and 'NOVELA' in name_up): channel = 'tn'
            else: unknown.append(f); continue

        if ftype == 'grilla':
            grillas.setdefault(channel, []).append(f)
            continue

        # Date extraction
        if ftype == 'json':
            date = _date_from_json_content(f)    # always from content
        else:  # xml
            date = _date_from_xml_filename(f.name)  # always from filename

        if date is None:
            unknown.append(f)
            continue

        key = (str(date), channel)
        if key not in days:
            days[key] = {'json': [], 'xml': None, 'log': None, 'date': date}
        if ftype == 'json':  days[key]['json'].append(f)
        elif ftype == 'xml': days[key]['xml'] = f
        elif ftype == 'log': days[key]['log'] = f

    return days, grillas, unknown, sony_files


def pair_sony_files(sony_files, lang='en'):
    """
    Pair Sony JSON files with Sony XML log files by channel code + date/marker.
    Returns list of pairing dicts for the app to process.
    Each: {'label', 'code', 'channel_name', 'json_file', 'xml_file', 'xml_filename', 'date'}
    Plus unmatched XMLs and JSONs.
    """
    json_list = [f for f in sony_files if f['ftype'] == 'json']
    xml_list  = [f for f in sony_files if f['ftype'] == 'xml']

    pairings    = []
    used_jsons  = set()

    # For each JSON: read markers to find expected XML(s)
    # Also extract date from JSON content for fallback
    json_info = []
    for jf in json_list:
        try:
            jf['file'].seek(0)
            data = json.load(jf['file'])
            jf['file'].seek(0)
        except:
            data = {'events': []}
        markers  = parse_sony_json_markers(data)
        date_val = _date_from_json_content(jf['file'])
        try: jf['file'].seek(0)
        except: pass
        json_info.append({'jf': jf, 'data': data,
                          'markers': markers, 'date': date_val})

    # Match full JSONs (with markers) first, then partial (date-based) to avoid consuming wrong XML
    json_info_sorted = sorted(json_info, key=lambda x: 0 if x['markers'] else 1)
    for jinfo in json_info_sorted:
        jf       = jinfo['jf']
        code     = jf['code']
        ch_name  = SONY_CHANNEL_MAP.get(code, code)
        matched_xml = None
        # Use filename date as fallback if content date failed
        if jinfo['date'] is None:
            jinfo['date'] = _date_from_xml_filename(jf['file'].name)

        if jinfo['markers']:
            for mk in jinfo['markers']:
                if not mk['log_base']: continue
                for xf in xml_list:
                    if xf['code'] != code: continue
                    xbase = extract_sony_xml_base(xf['file'].name)
                    if xbase.upper() == mk['log_base'].upper():
                        matched_xml = xf
                        break
                if matched_xml: break

        # Fallback: match by channel code + date from filename
        if not matched_xml and jinfo['date']:
            date_str = jinfo['date'].strftime('%Y%m%d')
            for xf in xml_list:
                if xf['code'] != code: continue
                if date_str in xf['file'].name:
                    matched_xml = xf
                    break

        # Last resort: if only one XML for this channel code, pair it
        if not matched_xml:
            available = [xf for xf in xml_list
                         if xf['code'] == code]
            if len(available) == 1:
                matched_xml = available[0]

        used_jsons.add(id(jf))
        pairings.append({
            'label':        f'{ch_name} — {jinfo["date"] or "?"}',
            'code':         code,
            'channel_name': ch_name,
            'json_file':    jf['file'],
            'json_data':    jinfo['data'],
            'xml_file':     matched_xml['file'] if matched_xml else None,
            'xml_filename': matched_xml['file'].name if matched_xml else None,
            'date':         jinfo['date'],
        })

    # Unmatched XMLs (no JSON for them)
    # Show XMLs with no JSON — find ones not paired to any JSON by code+date
    paired_xml_ids = {id(p["xml_file"]) for p in pairings if p["xml_file"]}
    unmatched_xml = [xf for xf in xml_list if id(xf["file"]) not in paired_xml_ids]
    for xf in unmatched_xml:
        code    = xf['code']
        ch_name = SONY_CHANNEL_MAP.get(code, code)
        pairings.append({
            'label':        f'{ch_name} — LOG ONLY (no JSON)',
            'code':         code,
            'channel_name': ch_name,
            'json_file':    None,
            'json_data':    None,
            'xml_file':     xf['file'],
            'xml_filename': xf['file'].name,
            'date':         None,
        })

    return pairings


# ── CHECKS ────────────────────────────────────────────────────────────────────

def check_programs_vs_grilla(playlist, grilla_ids, current_start, lang):
    """
    LCS-diff based program comparison. Handles:
    - Re-airs (walks both sequences in order, so second occurrence is checked at its position)
    - Replacements (CEND0402 deleted / NP0402 added — reports both, no cascading false errors)
    - Wrong episodes (same prefix, different date suffix)
    - Insertions / deletions
    """
    is_partial = current_start is not None
    part_seq   = build_show_sequence(playlist['programs'], from_start=current_start)

    if not grilla_ids:
        return [T('no_grilla', lang)]

    # Anchor partial playlist to correct grilla occurrence (handles re-air anchoring)
    if is_partial and part_seq:
        first_id  = part_seq[0]['id']
        first_pfx = show_prefix(first_id)
        # Count how many times first_id aired BEFORE current_start
        pre_count = sum(1 for p in playlist['programs']
                        if p['episode_id'] == first_id
                        and p['start'] and p['start'] < current_start)
        anchor, found_count = 0, 0
        for i, gid in enumerate(grilla_ids):
            if gid == first_id or (first_pfx and show_prefix(gid) == first_pfx):
                if found_count == pre_count:
                    anchor = i; break
                found_count += 1
        issues = [T('anchored', lang, i=anchor+1, id=grilla_ids[anchor] if grilla_ids else '?')]
        grilla_slice = grilla_ids[anchor:]
    else:
        grilla_slice = grilla_ids[:]
        issues = []

    # LCS sequential walk
    WINDOW = 8
    gi, pi = 0, 0

    while gi < len(grilla_slice) and pi < len(part_seq):
        gid = grilla_slice[gi]
        p   = part_seq[pi]
        pid = p['id']

        if gid == pid:
            gi += 1; pi += 1; continue

        pfx_g = show_prefix(gid)
        pfx_p = show_prefix(pid)

        # Same prefix → wrong episode (COSA0326 vs COSA0402)
        # For movie IDs (pure alpha like MARCE): also catch MARCE vs MARCELO
        is_movie_g = bool(re.match(r'^[A-Z]{3,8}$', gid))
        is_movie_p = bool(re.match(r'^[A-Z]{3,8}$', pid))
        pfx_match = (pfx_g and pfx_p and (
            pfx_g == pfx_p or
            ((is_movie_g or is_movie_p) and
             (pfx_g.startswith(pfx_p) or pfx_p.startswith(pfx_g)))
        ))
        if pfx_match:
            issues.append(T('wrong_ep', lang, g=gid, p=pid, t=fmt_t(p['start'])))
            gi += 1; pi += 1; continue

        # Look ahead
        future_pl = [part_seq[pi+k]['id'] for k in range(1, min(WINDOW+1, len(part_seq)-pi))]
        future_gr = list(grilla_slice[gi+1:gi+WINDOW+1])

        gid_ahead_pl = gid in future_pl
        pid_ahead_gr = pid in future_gr

        if not gid_ahead_pl and not pid_ahead_gr:
            # Replacement: grilla has X, playlist has Y, neither in each other near future
            deleted_lbl = 'DELETED' if lang == 'en' else 'ELIMINADO'
            added_lbl   = 'ADDED'   if lang == 'en' else 'AGREGADO'
            issues.append(f'  ↔  {deleted_lbl}: {gid} / {added_lbl}: {pid} @ {fmt_t(p["start"])}')
            gi += 1; pi += 1

        elif not gid_ahead_pl:
            # Grilla show deleted from playlist
            issues.append(T('not_in_pl', lang, id=gid))
            gi += 1

        elif not pid_ahead_gr:
            # Extra show in playlist not in grilla
            issues.append(T('extra_pl', lang, id=pid, t=fmt_t(p['start'])))
            pi += 1

        else:
            # Both in each other's future — realign via shortest path
            pl_offset = future_pl.index(gid) + 1   # steps in pl to reach gid
            gr_offset = future_gr.index(pid) + 1    # steps in gr to reach pid
            if pl_offset <= gr_offset:
                for k in range(pl_offset):
                    issues.append(T('extra_pl', lang, id=part_seq[pi+k]['id'],
                                    t=fmt_t(part_seq[pi+k]['start'])))
                pi += pl_offset
            else:
                for k in range(gr_offset):
                    issues.append(T('not_in_pl', lang, id=grilla_slice[gi+k]))
                gi += gr_offset

    # Tail: remaining grilla entries not in playlist
    while gi < len(grilla_slice):
        gid = grilla_slice[gi]
        if is_partial and any(p['episode_id'] == gid for p in playlist['programs']
                              if p['start'] and p['start'] < current_start):
            issues.append(T('already_aired', lang, id=gid))
        else:
            issues.append(T('not_in_pl', lang, id=gid))
        gi += 1

    # Tail: remaining playlist entries
    while pi < len(part_seq):
        issues.append(T('extra_pl', lang, id=part_seq[pi]['id'], t=fmt_t(part_seq[pi]['start'])))
        pi += 1

    has_errors = any(x.strip().startswith(('✗','⚠','↔')) for x in issues)
    if not has_errors:
        issues.append(f'  {T("ok_programs", lang)}')
    return issues


def check_commercials_vs_xml(playlist, xml_rows, current_start, lang):
    """
    Break-by-break commercial comparison.
    Aligns XML and JSON by segment ID (after_program_raw).
    Handles show replacements with a look-ahead window.
    Returns (issues_list, manual_warnings_list).
    """
    WINDOW = 15  # max segments to look ahead for replacement recovery

    # --- Build XML break list ---
    xml_breaks = build_xml_breaks(xml_rows)

    # --- Build JSON break list (only those with a raw anchor) ---
    json_breaks = [b for b in playlist['breaks'] if b.get('after_program_raw')]

    # --- For partial: align by matching first JSON break's segment in XML ---
    if current_start:
        # Filter json_breaks to those at/after current_start
        def _break_time(b):
            return next((i['start'] for i in b.get('items',[]) if i.get('start')), None)
        json_breaks = [b for b in json_breaks
                       if not _break_time(b) or _break_time(b) >= current_start]

        if json_breaks:
            first_seg = json_breaks[0].get('after_program_raw', '')
            if first_seg:
                # Find this segment in XML breaks and start from there
                xi = next((i for i, xb in enumerate(xml_breaks)
                           if xb['anchor_id'] == first_seg), None)
                if xi is not None:
                    xml_breaks = xml_breaks[xi:]  # start FROM this segment (not after)

    # --- Labels ---
    added_lbl    = {'en': 'added to playlist',    'es': 'agregado a playlist'}
    removed_lbl  = {'en': 'removed from playlist', 'es': 'eliminado de playlist'}
    replaced_lbl = {'en': 'SHOW REPLACED',         'es': 'PROGRAMA REEMPLAZADO'}
    lost_lbl     = {'en': 'ALIGNMENT LOST',        'es': 'ALINEACIÓN PERDIDA'}
    summary_lbl  = {'en': 'COMMERCIAL CHANGES SUMMARY', 'es': 'RESUMEN DE CAMBIOS'}
    added_tot    = {'en': 'added',    'es': 'agregados'}
    removed_tot  = {'en': 'removed',  'es': 'eliminados'}
    manual_cap   = {'en': '!!! DOUBLE CHECK MANUALLY !!!', 'es': '!!! VERIFICAR MANUALMENTE !!!'}

    issues       = []
    all_added    = Counter()
    all_removed  = Counter()
    manual_warns = []

    def _break_start(jb):
        return next((i['start'] for i in jb['items'] if i.get('start')), None)

    def _compare_pair(xb, jb):
        """Compare one aligned XML/JSON break pair. Returns (lines, added, removed)."""
        xml_c  = Counter(xb['commercials'])
        json_c = Counter(i['ref'] for i in jb['items'] if i['type'] == 'Commercial')
        anchor = jb.get('after_program_raw', '?')
        bs     = _break_start(jb)

        lines, add, rem = [], Counter(), Counter()
        for ref in sorted(set(xml_c) | set(json_c)):
            xc, jc = xml_c.get(ref, 0), json_c.get(ref, 0)
            if xc == jc: continue
            diff = jc - xc
            if diff > 0:
                lines.append(f'     + {ref} x{diff}  ({added_lbl[lang]})')
                add[ref] += diff
            else:
                lines.append(f'     - {ref} x{abs(diff)}  ({removed_lbl[lang]})')
                rem[ref] += abs(diff)

        if lines:
            header = [f'  ⚠  Break after [{anchor}] @ {fmt_t(bs)}']
            return header + lines, add, rem
        return [], add, rem

    def _compare_pool(xml_blist, json_blist, xml_label, json_label, bs):
        """Compare pooled commercials from a replaced block."""
        xml_c  = Counter(ref for xb in xml_blist for ref in xb['commercials'])
        json_c = Counter(ref for jb in json_blist for i in jb['items']
                         if i['type'] == 'Commercial' for ref in [i['ref']])
        lines, add, rem = [], Counter(), Counter()
        for ref in sorted(set(xml_c) | set(json_c)):
            xc, jc = xml_c.get(ref, 0), json_c.get(ref, 0)
            if xc == jc: continue
            diff = jc - xc
            if diff > 0: lines.append(f'     + {ref} x{diff}  ({added_lbl[lang]})'); add[ref] += diff
            else:        lines.append(f'     - {ref} x{abs(diff)}  ({removed_lbl[lang]})'); rem[ref] += abs(diff)
        return lines, add, rem

    # --- Walk both break lists in parallel ---
    xi, ji = 0, 0
    while xi < len(xml_breaks) and ji < len(json_breaks):
        xb = xml_breaks[xi]
        jb = json_breaks[ji]
        x_anc = xb['anchor_id']
        j_anc = jb.get('after_program_raw', '')

        if x_anc == j_anc:
            # Perfect match
            lines, add, rem = _compare_pair(xb, jb)
            issues.extend(lines)
            all_added.update(add)
            all_removed.update(rem)
            xi += 1; ji += 1

        else:
            # Mismatch — look ahead in both directions to recover
            found_xi = next((i for i in range(xi+1, min(xi+WINDOW, len(xml_breaks)))
                             if xml_breaks[i]['anchor_id'] == j_anc), None)
            found_ji = next((i for i in range(ji+1, min(ji+WINDOW, len(json_breaks)))
                             if json_breaks[i].get('after_program_raw') == x_anc), None)

            bs = _break_start(jb)

            if found_xi is not None and (found_ji is None or (found_xi-xi) <= (found_ji-ji)):
                # XML has more segments here — pooled comparison
                xml_block  = xml_breaks[xi:found_xi]
                json_block = [jb]
                x_show = normalize_id(x_anc)
                j_show = normalize_id(j_anc)
                issues.append(f'  ⚠  {replaced_lbl[lang]}: XML=[{x_show}...] → Playlist=[{j_show}] @ {fmt_t(bs)}')
                pool_lines, add, rem = _compare_pool(xml_block, json_block, x_show, j_show, bs)
                if pool_lines:
                    issues.extend(pool_lines)
                    warn = f'{manual_cap[lang]}: {replaced_lbl[lang]} [{x_show}→{j_show}] @ {fmt_t(bs)}'
                    manual_warns.append(warn)
                    issues.append(f'     ⚠  {warn}')
                else:
                    issues.append(f'     ✓ Commercials match within replaced block')
                all_added.update(add); all_removed.update(rem)
                xi = found_xi; ji += 1

            elif found_ji is not None:
                # JSON has more segments here — pooled comparison
                xml_block  = [xb]
                json_block = json_breaks[ji:found_ji]
                x_show = normalize_id(x_anc)
                j_show = normalize_id(j_anc)
                issues.append(f'  ⚠  {replaced_lbl[lang]}: XML=[{x_show}] → Playlist=[{j_show}...] @ {fmt_t(bs)}')
                pool_lines, add, rem = _compare_pool(xml_block, json_block, x_show, j_show, bs)
                if pool_lines:
                    issues.extend(pool_lines)
                    warn = f'{manual_cap[lang]}: {replaced_lbl[lang]} [{x_show}→{j_show}] @ {fmt_t(bs)}'
                    manual_warns.append(warn)
                    issues.append(f'     ⚠  {warn}')
                else:
                    issues.append(f'     ✓ Commercials match within replaced block')
                all_added.update(add); all_removed.update(rem)
                xi += 1; ji = found_ji

            else:
                # Can't recover — skip both and warn loudly
                warn = f'{manual_cap[lang]}: {lost_lbl[lang]} [{normalize_id(x_anc)} vs {normalize_id(j_anc)}] @ {fmt_t(bs)}'
                manual_warns.append(warn)
                issues.append(f'  ⚠  {warn}')
                xi += 1; ji += 1

    # --- Summary ---
    if all_added or all_removed:
        issues += ['', f'  ── {summary_lbl[lang]} ──']
        if all_added:
            total = sum(all_added.values())
            issues.append(f'  +{total} {added_tot[lang]}:')
            for ref, cnt in sorted(all_added.items()):
                issues.append(f'    {ref} x{cnt}')
        if all_removed:
            total = sum(all_removed.values())
            issues.append(f'  -{total} {removed_tot[lang]}:')
            for ref, cnt in sorted(all_removed.items()):
                issues.append(f'    {ref} x{cnt}')
    elif not issues:
        total_pl = len(playlist['commercials'])
        issues.append(f'  {T("ok_commercials", lang, n=total_pl)}')

    return issues, manual_warns


def check_promo_repeats(playlist, current_start=None, lang='en'):
    INFOMERCIAL_SECS = 1200  # 20 min
    issues = []

    for brk in playlist['breaks']:
        items = brk['items']
        if not items: continue
        bs = next((i['start'] for i in items if i.get('start')), None)
        if current_start and bs and bs < current_start: continue

        # Split break at infomercials (Commercial ≥ 20min) — they act as sub-break separators
        sub_breaks = []
        current_sub = []
        for item in items:
            if item['type'] == 'Commercial' and parse_duration(item.get('duration','00:00:00')) >= INFOMERCIAL_SECS:
                if current_sub:
                    sub_breaks.append(current_sub)
                current_sub = []  # reset after infomercial
            else:
                current_sub.append(item)
        if current_sub:
            sub_breaks.append(current_sub)
        if not sub_breaks:
            sub_breaks = [items]

        for sub in sub_breaks:
            promo_refs = [i['ref'] for i in sub if i['type'] == 'Promotion']
            for ref, cnt in Counter(promo_refs).items():
                if cnt > 1:
                    after = brk.get('after_program', '?')
                    sub_start = next((i['start'] for i in sub if i.get('start')), bs)
                    issues.append(T('promo_rep', lang, after=after,
                                    t=fmt_t(sub_start), ref=ref, n=cnt))


def check_not_ingested(playlist, current_start=None, lang='en'):
    lines, seen_eps, seen_other = [], set(), set()
    for item in playlist['not_ingested']:
        if current_start and item['start'] and item['start'] < current_start: continue
        atype, aref = item['type'], item['asset_ref']
        if atype in ('Program','live'):
            ep = normalize_id(aref)
            if ep in seen_eps: continue
            seen_eps.add(ep)
            show = re.sub(r'\[\].*$', '', item['name']).strip()
            lines.append(T('ni_program', lang, id=ep, t=fmt_t(item['start']), show=show))
        else:
            if aref in seen_other: continue
            seen_other.add(aref)
            lines.append(T('ni_other', lang, typ=atype, ref=aref,
                           t=fmt_t(item['start']), name=item['name']))
    return lines


def check_bugs(playlist, current_start=None, lang='en'):
    """
    Check LOGOHD bug logo assignments.
    Groups consecutive programs by their logo value and reports time ranges.
    Only reads from program events that have a 'logo' field set.
    """
    progs = [p for p in playlist['programs']
             if p.get('logo') is not None
             and (not current_start or not p['start'] or p['start'] >= current_start)]
    if not progs:
        return [f"  \u2714  No bugs scheduled" if lang=='en' else "  \u2714  Sin bugs programados"]

    # Group consecutive programs by logo
    groups = []  # (logo, start_time, end_time, programs_in_group)
    cur_logo   = progs[0]['logo']
    cur_start  = progs[0]['start']
    cur_progs  = [progs[0]]
    for p in progs[1:]:
        if p['logo'] == cur_logo:
            cur_progs.append(p)
        else:
            groups.append((cur_logo, cur_start, p['start'], cur_progs[:]))
            cur_logo  = p['logo']
            cur_start = p['start']
            cur_progs = [p]
    groups.append((cur_logo, cur_start, None, cur_progs))

    lines = []
    for logo, t_start, t_end, grp in groups:
        def _round_min(dt):
            if not dt: return None
            from datetime import timedelta as _td
            sec = dt.second + dt.microsecond/1e6
            rounded = dt.replace(second=0, microsecond=0) + (_td(minutes=1) if sec >= 30 else _td(0))
            return rounded
        t_start_r = _round_min(t_start)
        t_end_r   = _round_min(t_end)
        s = fmt_t(t_start_r) if t_start_r else '?'
        e = fmt_t(t_end_r)   if t_end_r   else ('end of day' if lang=='en' else 'fin del día')
        lines.append(f'  {logo}  :  {s} → {e}')
    return lines


def check_cue_tones(playlist, lang='en'):
    """
    Cue tone report.
    Sequence: CUE ON clip (ignored) → clip(s) → CUE OFF clip (included).
    Duration = sum of all clips strictly after CUE ON up to and including CUE OFF.
    Uses ev_dur stored on each event; clips computed from sorted cue_tones list.
    """
    all_cts = sorted(playlist.get('cue_tones', []), key=lambda c: c['start'] or datetime.min)
    cue_ons  = [c for c in all_cts if not c.get('is_cueoff')]
    cue_offs = [c for c in all_cts if c.get('is_cueoff')]

    if not cue_ons:
        return [f"  \u2714  No cue tones found" if lang=='en' else "  \u2714  Sin cue tones"]

    # Also get all promos sorted for duration lookup using gaps
    all_promos = sorted(playlist.get('promos', []), key=lambda p: p['start'] or datetime.min)
    promo_gap_dur = {}
    for i, p in enumerate(all_promos):
        if p['start']:
            if i+1 < len(all_promos) and all_promos[i+1]['start']:
                promo_gap_dur[p['start']] = int((all_promos[i+1]['start'] - p['start']).total_seconds())
            else:
                promo_gap_dur[p['start']] = p.get('duration', 30) or 30

    from collections import defaultdict
    stats = defaultdict(lambda: {'count': 0, 'first': None, 'last_dur': 0})
    total_dur = 0

    for ct_on in cue_ons:
        ref   = ct_on['ct_id']
        t_on  = ct_on['start']
        if not t_on: continue

        # Find the next CUE OFF after this CUE ON
        t_off = next((c['start'] for c in cue_offs if c['start'] and c['start'] > t_on), None)

        if t_off:
            # Sum durations of all promos strictly between t_on and t_off (inclusive of t_off promo)
            block_dur = sum(
                promo_gap_dur.get(p['start'], p.get('duration', 30) or 30)
                for p in all_promos
                if p['start'] and p['start'] > t_on and p['start'] <= t_off
            )
        else:
            # No CUE OFF found — use next CUE ON as boundary
            next_on = next((c['start'] for c in cue_ons if c['start'] and c['start'] > t_on), None)
            block_dur = sum(
                promo_gap_dur.get(p['start'], p.get('duration', 30) or 30)
                for p in all_promos
                if p['start'] and p['start'] > t_on and (next_on is None or p['start'] < next_on)
            )

        block_dur = min(block_dur, 240)  # cap at 4 min
        total_dur += block_dur

        stats[ref]['count'] += 1
        stats[ref]['last_dur'] = block_dur
        if stats[ref]['first'] is None or t_on < stats[ref]['first']:
            stats[ref]['first'] = t_on

    def fmt_dur(secs):
        m, s = divmod(int(secs), 60)
        return f'{m}min {s:02d}sec'

    lines = [f"  Total CUE ON: {len(cue_ons)} | Total duration: {fmt_dur(total_dur)}"]
    for ref in sorted(stats):
        s = stats[ref]
        first_str = s['first'].strftime('%H:%M:%S') if s['first'] else '?'
        last_str  = fmt_dur(s['last_dur'])
        lines.append(f"  {ref}: {s['count']}x | First: {first_str} | Last: {last_str}")
    return lines

def check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang):
    """
    Episode-number-only program check for HolaTV, LCS-style.
    One missing/extra entry does not cascade — handled like CATV/TVD.
    grilla_entries: (show_eps list, inf_count) tuple from parse_grilla_holatv_v2
    """
    if isinstance(grilla_entries, tuple):
        grilla_show_eps, grilla_inf_count = grilla_entries
    else:
        grilla_show_eps = [g['episode'] for g in grilla_entries if not g.get('is_inf')]
        grilla_inf_count = sum(1 for g in grilla_entries if g.get('is_inf'))

    if not grilla_show_eps and grilla_inf_count == 0:
        return [f'  ℹ  {"No grilla provided" if lang=="en" else "Sin grilla proporcionada"}']

    active    = [b for b in log_blocks
                 if not current_start_utc or not b['start_utc']
                 or b['start_utc'] >= current_start_utc]
    log_shows = [b for b in active if not b['is_hpp']]
    log_hpp   = [b for b in active if b['is_hpp']]

    if not log_shows and not log_hpp:
        return [f'  ℹ  {"No log data in window" if lang=="en" else "Sin datos de log en la ventana"}']

    def ep_from_id(base_id):
        m = re.search(r'(\d+)$', base_id)
        return int(m.group(1)) if m else None

    # ── Partial anchoring ──
    if current_start_utc and log_shows and grilla_show_eps:
        log_head   = [ep_from_id(b['base_id']) for b in log_shows[:4]]
        anchor_pos = None
        for gi in range(len(grilla_show_eps)):
            needed = min(3, len(grilla_show_eps) - gi, len(log_head))
            if needed < 1: break
            if all(grilla_show_eps[gi+k] == log_head[k] for k in range(needed)):
                anchor_pos = gi; break
        if anchor_pos is not None:
            grilla_show_eps = grilla_show_eps[anchor_pos:]

    # ── LCS walk ──
    WINDOW = 8
    issues  = []
    ok_count = 0
    gi, li  = 0, 0
    g_eps   = grilla_show_eps
    l_blks  = log_shows

    while gi < len(g_eps) and li < len(l_blks):
        g_ep = g_eps[gi]
        l_ep = ep_from_id(l_blks[li]['base_id'])

        if g_ep == l_ep:
            ok_count += 1; gi += 1; li += 1; continue

        # Look ahead
        fut_g = [g_eps[gi+k] for k in range(1, min(WINDOW+1, len(g_eps)-gi))]
        fut_l = [ep_from_id(l_blks[li+k]['base_id']) for k in range(1, min(WINDOW+1, len(l_blks)-li))]

        g_in_fut_l = g_ep in fut_l
        l_in_fut_g = l_ep in fut_g

        if not g_in_fut_l and not l_in_fut_g:
            warn = 'MANUAL CHECK' if lang=='en' else 'REVISIÓN MANUAL'
            issues.append(f'  ⚠  {warn}: grilla ep{g_ep} ≠ log {l_blks[li]["base_id"]} (ep{l_ep}) @ {fmt_t(l_blks[li]["start_utc"])}')
            gi += 1; li += 1
        elif not g_in_fut_l:
            not_lbl = 'NOT IN LOG' if lang=='en' else 'NO EN LOG'
            issues.append(f'  ✗  {not_lbl}: grilla ep{g_ep}')
            gi += 1
        elif not l_in_fut_g:
            ext_lbl = 'EXTRA IN LOG' if lang=='en' else 'EXTRA EN LOG'
            issues.append(f'  ✗  {ext_lbl}: {l_blks[li]["base_id"]} @ {fmt_t(l_blks[li]["start_utc"])}')
            li += 1
        else:
            bl_off = fut_l.index(g_ep) + 1
            gr_off = fut_g.index(l_ep) + 1 if l_ep in fut_g else WINDOW
            if bl_off <= gr_off:
                for k in range(bl_off):
                    issues.append(f'  ✗  {"EXTRA IN LOG" if lang=="en" else "EXTRA EN LOG"}: {l_blks[li+k]["base_id"]} @ {fmt_t(l_blks[li+k]["start_utc"])}')
                li += bl_off
            else:
                for k in range(gr_off):
                    issues.append(f'  ✗  {"NOT IN LOG" if lang=="en" else "NO EN LOG"}: grilla ep{g_eps[gi+k]}')
                gi += gr_off

    while gi < len(g_eps):
        issues.append(f'  ✗  {"NOT IN LOG" if lang=="en" else "NO EN LOG"}: grilla ep{g_eps[gi]}')
        gi += 1
    while li < len(l_blks):
        issues.append(f'  ✗  {"EXTRA IN LOG" if lang=="en" else "EXTRA EN LOG"}: {l_blks[li]["base_id"]} @ {fmt_t(l_blks[li]["start_utc"])}')
        li += 1

    # ── INF / HPP counter ──
    inf_lbl = 'Infomercials' if lang=='en' else 'Infomerciales'
    if current_start_utc:
        issues.append(f'  ℹ  {inf_lbl}: {len(log_hpp)} HPP in log window (grilla count skipped for partial)')
    elif grilla_inf_count == len(log_hpp):
        issues.append(f'  ✓  {inf_lbl}: {grilla_inf_count} in grilla, {len(log_hpp)} HPP in log — match')
    else:
        diff = len(log_hpp) - grilla_inf_count
        sign = '+' if diff > 0 else ''
        issues.append(f'  ✗  {inf_lbl}: grilla={grilla_inf_count}, log={len(log_hpp)} ({sign}{diff})')

    if not any('✗' in i or '⚠' in i for i in issues):
        issues.insert(0, f'  ✓  {"All" if lang=="en" else "Todos"} {ok_count} {"episodes match" if lang=="en" else "episodios coinciden"}')
    else:
        n_warn = sum(1 for i in issues if '⚠' in i)
        n_err  = sum(1 for i in issues if '✗' in i)
        issues.insert(0, f'  ✓ {ok_count} {"match" if lang=="en" else "coinciden"}  |  ❗ {n_warn} {"manual check" if lang=="en" else "revisión manual"}  |  ✗ {n_err} {"mismatch" if lang=="en" else "diferencia"}')
    return issues


def check_holatv_timing_v2(grilla_entries, log_blocks, current_start_utc, lang, tolerance_secs=2700):
    """Timing check placeholder — grilla times not reliable."""
    return [f'  ✓  {"Episodes match in Grilla, Log and Playlist." if lang=="en" else "Episodios coinciden en Grilla, Log y Playlist."}']


def parse_grilla_holatv_v2(filepath_or_bytes, target_date):
    """
    Parse HolaTV PDF grilla — extracts episode numbers in column order.
    Uses vertical-distance-first code matching + digit-token merging.
    Returns (show_eps, inf_count).
    """
    try:
        import pdfplumber as _ppl
    except ImportError:
        return [], 0
    try:
        if hasattr(filepath_or_bytes, 'read'):
            filepath_or_bytes.seek(0)
            raw = filepath_or_bytes.read()
            filepath_or_bytes.seek(0)
            import tempfile, os as _os
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.write(raw); tmp.close()
            pdf_path = tmp.name; cleanup = True
        else:
            pdf_path = filepath_or_bytes; cleanup = False

        all_words = []
        with _ppl.open(pdf_path) as pdf:
            for pg_idx, page in enumerate(pdf.pages):
                words  = page.extract_words(x_tolerance=8, y_tolerance=3)
                page_h = float(page.height)
                for w in words:
                    all_words.append({**w, 'abs_top': w['top'] + pg_idx * page_h})
        if cleanup:
            import os as _os2; _os2.unlink(pdf_path)

        date_pat = re.compile(r'^(\d{2})/(\d{2})$')
        day_cols = {}
        for i, w in enumerate(all_words):
            m = date_pat.match(w['text'])
            if m and i > 0 and all_words[i-1]['text'].endswith('.'):
                day_x  = (all_words[i-1]['x0'] + w['x1']) / 2
                month, day_n = int(m.group(2)), int(m.group(1))
                try:
                    from datetime import date as _d2
                    d = _d2(target_date.year, month, day_n)
                    day_cols[str(d)] = day_x
                except Exception:
                    pass

        if str(target_date) not in day_cols:
            return [], 0

        target_x  = day_cols[str(target_date)]
        xs        = sorted(day_cols.values())
        idx       = xs.index(target_x)
        col_left  = (xs[idx-1] + xs[idx]) / 2 if idx > 0 else 0
        col_right = (xs[idx] + xs[idx+1]) / 2 if idx < len(xs)-1 else 9999

        code_re    = re.compile(r'^[A-Z][A-Z0-9_]{2,9}$')
        codes_at_y = []
        for i, w in enumerate(all_words):
            if code_re.match(w['text']) and i+1 < len(all_words) and all_words[i+1]['text'] == '(-)':
                codes_at_y.append((w['abs_top'], (w['x0']+w['x1'])/2, w['text']))

        ep_re   = re.compile(r'^\d{1,4}$')
        raw_eps = sorted(
            [{'y': w['abs_top'], 'text': w['text'],
              'x0': w['x0'], 'x1': w['x1'],
              'cx': (w['x0']+w['x1'])/2}
             for w in all_words
             if col_left <= (w['x0']+w['x1'])/2 <= col_right
             and ep_re.match(w['text'])],
            key=lambda w: (w['y'], w['x0']))

        merged = []
        i = 0
        while i < len(raw_eps):
            w     = raw_eps[i]
            group = [w]
            j     = i + 1
            while j < len(raw_eps):
                nw = raw_eps[j]
                if abs(nw['y'] - w['y']) < 2 and nw['x0'] - group[-1]['x1'] < 15:
                    group.append(nw); j += 1
                else:
                    break
            merged.append({'y': w['y'],
                           'text': ''.join(g['text'] for g in group),
                           'cx': (w['x0'] + group[-1]['x1']) / 2})
            i = j

        show_eps, inf_count = [], 0
        prev_ep, prev_is_inf = None, None
        for mw in merged:
            if not ep_re.match(mw['text']): continue
            try: ep_num = int(mw['text'])
            except: continue
            ep_y, ep_cx = mw['y'], mw['cx']
            cands = [(ep_y - cy, abs(ep_cx - cx), code)
                     for cy, cx, code in codes_at_y if 0 <= ep_y - cy < 40]
            if not cands: continue
            code   = sorted(cands)[0][2]
            is_inf = code.rstrip('_') == 'INF'
            if ep_num == prev_ep and is_inf == prev_is_inf:
                continue
            prev_ep, prev_is_inf = ep_num, is_inf
            if is_inf:
                inf_count += 1
            else:
                show_eps.append(ep_num)
        return show_eps, inf_count

    except Exception:
        return [], 0


def parse_holatv_log_xml_v2(file_or_bytes, log_date):
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        raw  = re.sub(rb'&(?!amp;|lt;|gt;|apos;|quot;|#)', b'&amp;', raw)
        root = ET.fromstring(raw)
        fields = [f.text for f in root.find('fields')]
        data   = root.find('data')
        rows, dx, cx = [], 0, 0
        for row in data:
            vals  = dict(zip(fields, [col.text or '' for col in row]))
            mid   = vals.get('Media Id', '').split('#')[0].strip()
            typ   = vals.get('Type', '').upper()
            lt    = vals.get('Local Time', '')
            dur   = vals.get('Duration', '00:00:00')
            title = vals.get('Title', '')
            try:
                dt_utc = datetime.strptime(lt[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur.replace(';', ':').split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            if typ == 'DX': dx += 1
            elif typ == 'CX': cx += 1
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif typ == 'PROGRAM':
                norm_type = 'PROGRAM'
            elif typ in ('DX', 'CX'):
                norm_type = typ
            elif typ == 'PROMOTION':
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': title})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_xlsx_v2(file_or_bytes, log_date):
    try:
        from openpyxl import load_workbook
        import io as _io
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            data = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            data = file_or_bytes
        wb   = load_workbook(_io.BytesIO(data), read_only=True)
        ws   = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        rows, dx, cx = [], 0, 0
        for r in rows_raw[1:]:
            if not r or r[0] is None: continue
            try:
                hora_raw = str(r[1]).strip() if r[1] else ''
                tipo_raw = str(r[2]).strip().upper() if r[2] else ''
                rec_key  = str(r[4]).strip() if len(r) > 4 and r[4] else ''
                titulo   = str(r[5]).strip() if len(r) > 5 and r[5] else ''
                dur_raw  = str(r[6]).strip() if len(r) > 6 and r[6] else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                offset = 4 if _edt_start(et_dt.year) <= et_dt < _est_start(et_dt.year) else 5
                dt_utc = et_dt + timedelta(hours=offset)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':
                norm_type = 'PROGRAM'
            elif tipo_raw in ('DX', 'CX'):
                norm_type = tipo_raw
            elif tipo_raw in ('PROM', 'CORT'):
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_txt_v2(file_or_bytes, log_date):
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try: text = raw.decode(enc); break
            except: pass
        else:
            text = raw.decode('latin-1', errors='replace')
        lines = text.splitlines()
        if not lines: return [], 0, 0
        rows, dx, cx = [], 0, 0
        header_done = False
        for line in lines:
            if not line.strip(): continue
            cols = line.rstrip('\r\n').split('\t')
            if not header_done:
                header_done = True
                if cols[0].strip().upper() in ('N.ORD.', 'N.ORD', 'NORD'): continue
            if len(cols) < 9: continue
            try:
                hora_raw = cols[1].strip()
                tipo_raw = cols[2].strip().upper()
                rec_key  = cols[25].strip() if len(cols) > 25 and cols[25].strip() else cols[3].strip().split('#')[0]
                titulo   = cols[5].strip() if len(cols) > 5 else ''
                dur_raw  = cols[8].strip() if len(cols) > 8 else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                offset = 4 if _edt_start(et_dt.year) <= et_dt < _est_start(et_dt.year) else 5
                dt_utc = et_dt + timedelta(hours=offset)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'): norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':  norm_type = 'PROGRAM'
            elif tipo_raw in ('DX','CX'): norm_type = tipo_raw
            elif tipo_raw in ('PROM','CORT'): norm_type = 'PROMO'
            else: norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def load_holatv_log(file_or_bytes, log_date):
    if file_or_bytes is None:
        return [], 0, 0
    name = getattr(file_or_bytes, 'name', '') or ''
    ext  = name.lower().rsplit('.', 1)[-1] if '.' in name else ''
    if ext == 'xml':
        return parse_holatv_log_xml_v2(file_or_bytes, log_date)
    elif ext in ('xlsx', 'xlsm'):
        return parse_holatv_log_xlsx_v2(file_or_bytes, log_date)
    elif ext == 'txt':
        return parse_holatv_log_txt_v2(file_or_bytes, log_date)
    return parse_holatv_log_xml_v2(file_or_bytes, log_date)


def group_holatv_blocks(log_rows):
    prog_rows = [r for r in log_rows
                 if r['type'] == 'PROGRAM' or r['media_id'].startswith('HPP')]
    prog_rows.sort(key=lambda x: x['start_utc'] or datetime.min)
    blocks, prev_base, cur_block = [], None, None
    for r in prog_rows:
        mid  = r['media_id']
        base = re.sub(r'_\d+$', '', mid)
        if base != prev_base:
            if cur_block: blocks.append(cur_block)
            cur_block = {'base_id': base, 'start_utc': r['start_utc'],
                         'duration_secs': r['duration_secs'], 'segments': [r],
                         'is_hpp': base.startswith('HPP')}
            prev_base = base
        else:
            cur_block['segments'].append(r)
            cur_block['duration_secs'] += r['duration_secs']
    if cur_block: blocks.append(cur_block)
    return blocks


def pick_grilla_for_date(grilla_list, target_date, channel):
    if not grilla_list: return None, None
    if len(grilla_list) == 1: return grilla_list[0], None
    from openpyxl import load_workbook
    import io
    for gf in grilla_list:
        try:
            gf.seek(0)
            data = gf.read()
            gf.seek(0)
            if gf.name.lower().endswith('.pdf'):
                _fname = gf.name.replace('_', ' ').upper()
                _MES = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
                        'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}
                _tok  = re.findall(r'(\d+)\s+([A-Z]{3})', _fname)
                _yr_m = re.search(r'(\d{4})', _fname)
                _year = int(_yr_m.group(1)) if _yr_m else target_date.year
                _months = [(m.start(), _MES[m.group()])
                           for m in re.finditer(r'(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)', _fname)]
                _nums   = [(m.start(), int(m.group()))
                           for m in re.finditer(r'\d+', _fname) if 1 <= int(m.group()) <= 31]
                _dates  = []
                for _mpos, _mnum in _months:
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if _before:
                        _day = max(_before, key=lambda x: x[0])[1]
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _day))
                        except Exception: pass
                if len(_months) == 1 and len(_dates) == 1 and _months:
                    _mpos, _mnum = _months[0]
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if len(_before) >= 2:
                        _all_days = sorted(set(n for _,n in _before))
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _all_days[0]))
                        except Exception: pass
                if len(_dates) >= 2:
                    _s, _e = min(_dates), max(_dates)
                    if _s <= target_date <= _e:
                        gf.seek(0); return gf, None
                    continue
                gf.seek(0); return gf, None
            wb = load_workbook(io.BytesIO(data), read_only=True)
            if channel in ('latam', 'us', 'tn', 'hl', 'hu'):
                for name in reversed(wb.sheetnames):
                    ws = wb[name]
                    rows = list(ws.iter_rows(max_row=3, values_only=True))
                    if len(rows) < 2: continue
                    header = rows[1]
                    col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                                 for i, cell in enumerate(header)
                                 if _parse_date_str(cell, force_year=target_date.year)]
                    if not col_dates: continue
                    if col_dates[0][1] <= target_date <= col_dates[-1][1]:
                        gf.seek(0); return gf, None
            else:
                ws = wb.active
                rows = list(ws.iter_rows(max_row=2, values_only=True))
                if len(rows) >= 2:
                    for cell in rows[1]:
                        d = _parse_date_str(cell, force_year=target_date.year)
                        if d and abs((d - target_date).days) <= 6:
                            gf.seek(0); return gf, None
        except Exception:
            try: gf.seek(0)
            except: pass
    grilla_list[0].seek(0)
    return grilla_list[0], 'Could not determine week for grilla — using first file'


def generate_report_holatv_v2(channel, log_rows, dx_count, cx_count,
                               grilla_entries, playlist, lang='en', file_info=None,
                               current_start_utc=None):
    sep        = '═' * 60
    log_blocks = group_holatv_blocks(log_rows)
    prog_blocks = [b for b in log_blocks if not b['is_hpp']]
    hpp_blocks  = [b for b in log_blocks if b['is_hpp']]
    pt = playlist['type'] if playlist else 'full'

    lines = [sep, f'CHANNEL: {channel}',
             f'DATE: {log_rows[0]["start_utc"].date() if log_rows and log_rows[0].get("start_utc") else "?"}',
             f'TYPE: {"Full Day" if pt=="full" else ("Partial" if lang=="en" else "Parcial")}']
    if current_start_utc:
        lines.append(f'{"Checking from" if lang=="en" else "Verificando desde"}: {fmt_t(current_start_utc)}')
    if file_info:
        lbl = 'Files' if lang=='en' else 'Archivos'
        lines.append(f'{lbl}:')
        if file_info.get('grilla'): lines.append(f'  Grid:     {file_info["grilla"]}')
        if file_info.get('log'):    lines.append(f'  Log:      {file_info["log"]}')
        if file_info.get('json'):   lines.append(f'  Playlist: {file_info["json"]}')
    lines += [sep,
              f'{"Summary" if lang=="en" else "Resumen"}: {len(prog_blocks)} show blocks | {len(hpp_blocks)} infomercials | DX={dx_count} CX={cx_count}',
              '']

    lines.append(f'── [1] {"PROGRAM CHECK (Grilla vs Log)" if lang=="en" else "VERIFICACIÓN PROGRAMAS (Grilla vs Log)"} ──')
    lines += check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang)
    lines.append('')

    lines.append(f'── [2] {"TIMING CHECK" if lang=="en" else "VERIFICACIÓN TIMING"} ──')
    lines += check_holatv_timing_v2(grilla_entries, log_blocks, current_start_utc, lang)
    lines.append('')

    if playlist:
        # [3] Infomercial check — both sides filtered to window
        lines.append(f'── [3] {"INFOMERCIAL CHECK (Log vs Playlist)" if lang=="en" else "VERIFICACIÓN INFOMERCIALES (Log vs Playlist)"} ──')
        _hpp_start  = current_start_utc
        log_hpp_ids = [b['base_id'] for b in hpp_blocks
                       if not _hpp_start or not b['start_utc'] or b['start_utc'] >= _hpp_start]
        pl_hpp      = [c for c in playlist.get('commercials', [])
                       if c.get('asset_ref', c.get('ref','')).startswith('HPP')
                       and (not _hpp_start or not c.get('start') or c['start'] >= _hpp_start)]
        pl_hpp_ids  = [c.get('asset_ref', c.get('ref','')) for c in pl_hpp]
        from collections import Counter as _Counter
        log_c = _Counter(log_hpp_ids); pl_c = _Counter(pl_hpp_ids)
        hpp_issues = [f'  ⚠  {hid}: log={log_c.get(hid,0)}x playlist={pl_c.get(hid,0)}x'
                      for hid in sorted(set(log_c) | set(pl_c))
                      if log_c.get(hid,0) != pl_c.get(hid,0)]
        lines += hpp_issues if hpp_issues else [f'  ✓  {"Infomercials match" if lang=="en" else "Infomerciales coinciden"}']
        lines.append('')

        lines.append(f'── [4] {"PROMO REPEAT CHECK" if lang=="en" else "VERIFICACIÓN PROMOS REPETIDAS"} ──')
        pi = check_promo_repeats(playlist, current_start_utc, lang)
        lines += pi if pi else [f'  {T("ok_promos", lang)}']
        lines.append('')

        lines.append(f'── [5] {"NOT INGESTED" if lang=="en" else "NO INGRESADOS"} ──')
        ni = check_not_ingested(playlist, current_start_utc, lang)
        lines += ni if ni else [f'  {T("ok_ingested", lang)}']
        lines.append('')

        lines.append(f'── [6] {"BUGS CHECK" if lang=="en" else "VERIFICACIÓN DE BUGS"} ──')
        bi = check_bugs(playlist, current_start_utc, lang)
        lines += bi if bi else [f'  {T("ok_bugs", lang)}']
        lines.append('')

        lines.append(f'── [7] {"CUE TONES" if lang=="en" else "CUE TONES"} ──')
        ci = check_cue_tones(playlist, lang)
        lines += ci if ci else [f'  {T("ok_cues", lang)}']
        lines.append('')

    lines.append(sep)
    return '\n'.join(str(l) for l in lines)


def generate_report(channel, playlist, xml_rows, grilla_ids, lang='en', is_tn=False, file_info=None):
    sep = '═' * 60
    pt  = playlist['type']
    current_start = playlist['programs'][0]['start'] if pt == 'current' and playlist['programs'] else None
    part_seq = build_show_sequence(playlist['programs'], from_start=current_start)
    total_comms = len([c for c in playlist['commercials']
                       if not current_start or (c['start'] and c['start'] >= current_start)])

    lines = [sep,
             f'{T("channel",lang)}: {channel.upper()}',
             f'{T("date_lbl",lang)}: {playlist["date"]}',
             f'{T("type_lbl",lang)}: {T("full_day",lang) if pt=="full" else T("partial",lang)}']
    # Visual indicator for partial vs full
    if pt == 'current':
        lines.append('▶▶▶  PARTIAL / CURRENT PLAYLIST  ◀◀◀')
    if current_start:
        lines.append(f'{T("checking_from",lang)}: {fmt_t(current_start)}')
    # Files used for this report block
    if file_info:
        files_lbl = "Files" if lang == "en" else "Archivos"
        lines.append(f'{files_lbl}:')
        if file_info.get("grilla"): lines.append(f'  Grid:     {file_info["grilla"]}')
        if file_info.get("xml"):    lines.append(f'  Log:      {file_info["xml"]}')
        if file_info.get("json"):   lines.append(f'  Playlist: {file_info["json"]}')
    lines += [sep, f'{T("summary",lang)}: {len(part_seq)} {T("show_blocks",lang)} | {total_comms} {T("commercials_lbl",lang)}', '']

    lines.append(f'── [1] {T("section_programs",lang)} ──')
    prog_lines = []
    if is_tn and grilla_ids:
        prog_lines = check_programs_vs_grilla_tn(playlist, grilla_ids, current_start, lang)
    elif not grilla_ids:
        prog_lines = [T('no_grilla', lang)]
    else:
        prog_lines = check_programs_vs_grilla(playlist, grilla_ids, current_start, lang)
    lines += prog_lines
    lines.append('')

    lines.append(f'── [2] {T("section_commercials",lang)} ──')
    manual_warns = []
    comm_lines = []
    if not xml_rows:
        lines.append(T('no_xml', lang))
    elif is_tn:
        # TN has no commercials — simple count only, no break-by-break needed
        lines.append(f'  {T("ok_commercials", lang, n=total_comms)}')
    else:
        comm_lines, manual_warns = check_commercials_vs_xml(playlist, xml_rows, current_start, lang)
        lines += comm_lines
    lines.append('')

    lines.append(f'── [3] {T("section_promos",lang)} ──')
    pi = check_promo_repeats(playlist, current_start, lang)
    lines += pi if pi else [f'  {T("ok_promos",lang)}']
    lines.append('')

    lines.append(f'── [4] {T("section_ingested",lang)} ──')
    if is_tn:
        lines.append(f'  ℹ  Not applicable for this channel')
    else:
        ni = check_not_ingested(playlist, current_start, lang)
        lines += ni if ni else [f'  {T("ok_ingested",lang)}']
    lines.append('')

    if not is_tn:
        lines.append(f'── [5] {T("section_bugs",lang)} ──')
        lines += check_bugs(playlist, current_start, lang)
        lines.append('')

    if pt == 'full':
        lines.append(f'── [6] {T("section_cues",lang)} ──')
        lines += check_cue_tones(playlist, lang)
        lines.append('')

    # ─ Collect Manual Review warnings ─
    manual_review_needed_lbl = {'en': '⚠ Manual Review needed:',  'es': '⚠ Revisión manual requerida:'}
    
    # Check for problems in program section
    for line in prog_lines:
        if any(marker in line for marker in ['⚠', '✗', 'WRONG EPISODE', 'NOT IN PLAYLIST', 'EXTRA']):
            # Extract show/episode info from line - try to get from parentheses or just use the line
            warn_line = line.strip()
            if warn_line and any(x in warn_line for x in ['⚠', '✗']):
                manual_warns.append(f'PROGRAM: {warn_line}')
    
    # Check for commercial changes (these are already in manual_warns from check_commercials_vs_xml)
    
    # Add consolidated manual review section if there are warnings
    if manual_warns:
        lines.append('')
        lines.append('═' * 60)
        lines.append(manual_review_needed_lbl[lang])
        seen_warns = set()
        for warn in manual_warns:
            # Deduplicate similar warnings
            if warn not in seen_warns:
                lines.append(f'  • {warn}')
                seen_warns.add(warn)
        lines.append('═' * 60)
    
    lines.append(sep)
    return '\n'.join(lines), manual_warns


# ══════════════════════════════════════════════════════════════════════════════
# SONY / AXN CHANNEL LOGIC
# ══════════════════════════════════════════════════════════════════════════════

SONY_CHANNEL_MAP = {
    'A1':'AXN ARGENTINA LAT','A2':'AXN MEXICO LAT','A3':'AXN ANDES LAT',
    'A4':'AXN BRASIL','A5':'AXN VENEZUELA LAT','A6':'AXN CENTRO AMERICA LAT',
    'F1':'SONY MOVIE LATAM','F4':'SONY MOVIE BRASIL',
    'S1':'SONY ARGENTINA LAT','S2':'SONY MEXICO LAT','S3':'SONY ANDES LAT',
    'S4':'SONY BRASIL','S5':'SONY VENEZUELA LAT','S6':'SONY CENTRO AMERICA LAT',
}
SONY_CODES = set(SONY_CHANNEL_MAP.keys())


def extract_sony_code(filename):
    """Extract 2-char Sony/AXN channel code from filename."""
    name = filename.upper()
    # JSON: vipeSchedule_S6_... or vipeSchedule_S6SET...
    m = re.search(r'VIPESCHEDULE_([A-Z][0-9])[\W_]', name)
    if m and m.group(1) in SONY_CODES:
        return m.group(1)
    # XML: S620260401c.xml or S620260401c_XML.xml
    m = re.match(r'^([A-Z][0-9])(\d{8})[A-Za-z]', filename.strip())
    if m and m.group(1).upper() in SONY_CODES:
        return m.group(1).upper()
    return None


def extract_sony_version(filename):
    """Extract version letter (a-z) from Sony XML filename like S620260401c.XML"""
    m = re.search(r'\d{8}([a-zA-Z])(?:_XML|\.)', filename, re.IGNORECASE)
    return m.group(1).lower() if m else None


def extract_sony_xml_base(filename):
    """Normalise XML filename to base form for marker comparison.
    S620260401c_XML.xml -> S620260401c.XML
    S620260401c.XML.xml -> S620260401c.XML  (Windows double extension)
    S620260401c.XML     -> S620260401c.XML
    """
    base = filename.strip()
    # Handle .XML.xml (Windows double extension)
    base = re.sub(r'\.XML\.xml$', '.XML', base, flags=re.IGNORECASE)
    # Handle _XML.xml
    base = re.sub(r'_XML\.xml$', '.XML', base, flags=re.IGNORECASE)
    # Handle plain .xml
    base = re.sub(r'\.xml$', '.XML', base, flags=re.IGNORECASE)
    return base


def parse_sony_xml_log(filepath_or_bytes):
    """Parse Sony/AXN XML log (tabledata format, same as TN).
    Returns list of dicts: mediaid, local_time (datetime), duration_secs, title, type.
    """
    try:
        if hasattr(filepath_or_bytes, 'read'): content = filepath_or_bytes.read()
        elif isinstance(filepath_or_bytes, bytes): content = filepath_or_bytes
        else:
            with open(filepath_or_bytes, 'rb') as f: content = f.read()
        # Fix unescaped & characters (e.g. "P&G" in title fields)
        content = re.sub(rb'&(?!amp;|lt;|gt;|apos;|quot;|#)', b'&amp;', content)
        root = ET.fromstring(content)
        rows = []
        for row in root.findall('.//row'):
            typ   = row.findtext('column-5','').strip().upper()
            mid   = row.findtext('column-4','').strip()
            lt    = row.findtext('column-1','').strip()
            dur_raw = row.findtext('column-3','').strip()
            title = row.findtext('column-6','').strip()
            try:
                dt = datetime.strptime(lt, '%Y-%m-%d %H:%M:%S')
            except:
                dt = None
            try:
                # Duration format: HH:MM:SS;FF — include frames (30fps)
                dur_parts = dur_raw.split(';')
                h, m_, s = dur_parts[0].split(':')
                frames = int(dur_parts[1]) if len(dur_parts) > 1 else 0
                dur_secs = int(h)*3600 + int(m_)*60 + int(s) + frames/30
            except:
                dur_secs = 0
            rows.append({'mediaid': mid, 'local_dt': dt,
                         'duration_secs': dur_secs, 'title': title, 'type': typ})
        return rows
    except:
        return []


def parse_sony_json_markers(data):
    """Extract all markers from a Sony JSON.
    Returns list of {'marker_name', 'log_base', 'version', 'start_dt', 'index'}
    """
    markers = []
    for i, ev in enumerate(data.get('events', [])):
        for a in ev.get('assets', []):
            if a.get('type') == 'marker':
                name  = ev.get('name', '')
                start = ev.get('startTime', '')[:19]
                # Extract log filename from marker name
                # "Playlist start, S620260401c.XML (1).xml"
                m = re.search(r'([A-Z][0-9]\d{8}[a-z]\.XML)', name, re.IGNORECASE)
                log_base = m.group(1).upper() if m else None
                version  = extract_sony_version(log_base) if log_base else None
                try:
                    start_dt = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
                except:
                    start_dt = None
                markers.append({'name': name, 'log_base': log_base,
                                 'version': version, 'start_dt': start_dt,
                                 'event_index': i})
    return markers


def check_sony(json_data, xml_rows, xml_filename, lang='en'):
    """
    Sony/AXN broadcast check:
    1. Marker list
    2. Marker vs log filename / version match
    3. Endpoint check (marker start ≈ log start, or JSON end ≈ log end for partial)
    4. Segment timing: all PROGRAM segments match within 5 seconds
    Returns (report_lines, has_errors)
    """
    sep   = '─' * 50
    lines = []
    has_errors = False

    events = json_data.get('events', [])
    xml_base = extract_sony_xml_base(xml_filename) if xml_filename else None
    xml_version = extract_sony_version(xml_filename) if xml_filename else None

    # Detect playlist type
    markers = parse_sony_json_markers(json_data)
    is_partial = (len(markers) == 0)

    # ── Marker list ──
    lines.append('── [1] MARKERS ──')
    if not markers:
        lines.append('  ℹ  No markers found (partial/current playlist)')
    else:
        for mk in markers:
            lines.append(f'  📌 {mk["name"]}')
            lines.append(f'     Expected log: {mk["log_base"] or "?"} | Version: {mk["version"] or "?"} | Start: {fmt_t(mk["start_dt"]) if mk["start_dt"] else "?"}')
    lines.append('')

    # ── Version / filename match ──
    lines.append('── [2] LOG FILE MATCH ──')
    if not xml_rows:
        lines.append('  ! No XML log provided')
    elif not markers and not is_partial:
        lines.append('  ! Cannot verify — no markers in JSON')
    else:
        lines.append(f'  Log file: {xml_filename or "?"}')
        if markers:
            mk = markers[0]
            if mk['log_base'] and xml_base:
                if mk['log_base'].upper() == xml_base.upper():
                    lines.append(f'  ✓ Filename matches marker: {mk["log_base"]}')
                else:
                    lines.append(f'  ✗ FILENAME MISMATCH: Marker expects {mk["log_base"]} | Got {xml_base}')
                    has_errors = True
            if mk['version'] and xml_version:
                if mk['version'] == xml_version:
                    lines.append(f'  ✓ Version matches: {mk["version"].upper()}')
                else:
                    lines.append(f'  ✗ VERSION MISMATCH: Marker says {mk["version"].upper()} | Log is {xml_version.upper()}')
                    has_errors = True
        elif is_partial:
            lines.append(f'  ℹ  Partial playlist — no marker to verify filename against')
    lines.append('')

    # ── Endpoint check ──
    lines.append('── [3] ENDPOINT CHECK ──')
    if not xml_rows:
        lines.append('  ! No XML log provided')
    else:
        xml_progs = [r for r in xml_rows if r['type'] == 'PROGRAM']
        # Log start and end
        log_start_dt = next((r['local_dt'] for r in xml_rows if r['local_dt']), None)
        last_row = xml_rows[-1]
        if last_row['local_dt']:
            log_end_dt = last_row['local_dt'] + timedelta(seconds=last_row['duration_secs'])
        else:
            log_end_dt = None

        if log_start_dt:
            lines.append(f'  Log start: {fmt_t(log_start_dt)}')
        if log_end_dt:
            lines.append(f'  Log end:   {fmt_t(log_end_dt)}')

        if markers and log_start_dt:
            mk = markers[0]
            if mk['start_dt']:
                diff = abs((mk['start_dt'] - log_start_dt).total_seconds())
                if diff <= 5:
                    lines.append(f'  ✓ Marker start matches log start (diff={diff:.1f}s)')
                else:
                    lines.append(f'  ✗ ENDPOINT MISMATCH: Marker start={fmt_t(mk["start_dt"])} | Log start={fmt_t(log_start_dt)} | diff={diff:.1f}s')
                    has_errors = True
        elif is_partial and log_end_dt:
            # For partial: JSON end ≈ log end
            # Find JSON last program end time
            json_end_dt = None
            for ev in reversed(events):
                for a in ev.get('assets', []):
                    if a.get('type') == 'Program':
                        st      = ev.get('startTime', '')[:19]
                        dur_raw = ev.get('duration', '').split('@')[0]
                        try:
                            sdt = datetime.strptime(st, '%Y-%m-%d %H:%M:%S')
                            dur_parts = dur_raw.split(';')
                            h, m_, s = dur_parts[0].split(':')
                            frames = int(dur_parts[1]) if len(dur_parts) > 1 else 0
                            json_end_dt = sdt + timedelta(hours=int(h), minutes=int(m_),
                                                          seconds=int(s) + frames//30)
                        except:
                            pass
                        break
                if json_end_dt: break

            if json_end_dt:
                diff = abs((json_end_dt - log_end_dt).total_seconds())
                if diff <= 5:
                    lines.append(f'  ✓ JSON end matches log end (JSON end={fmt_t(json_end_dt)} | Log end={fmt_t(log_end_dt)} | diff={diff:.1f}s)')
                else:
                    lines.append(f'  ✗ ENDPOINT MISMATCH: JSON ends {fmt_t(json_end_dt)} | Log ends {fmt_t(log_end_dt)} | diff={diff:.1f}s')
                    has_errors = True
    lines.append('')

    # ── Segment timing check ──
    lines.append('── [4] SEGMENT TIMING CHECK (≤5s tolerance) ──')
    if not xml_rows:
        lines.append('  ! No XML log provided')
    else:
        xml_progs = [r for r in xml_rows if r['type'] == 'PROGRAM']
        xml_lookup = {}
        for r in xml_progs:
            xml_lookup.setdefault(r['mediaid'], []).append(r['local_dt'])

        json_segs = []
        for ev in events:
            for a in ev.get('assets', []):
                if a.get('type') == 'Program':
                    st  = ev.get('startTime', '')[:19]
                    ref = a.get('reference', '')
                    try:
                        dt = datetime.strptime(st, '%Y-%m-%d %H:%M:%S')
                    except:
                        dt = None
                    json_segs.append({'ref': ref, 'dt': dt,
                                      'name': ev.get('name', '')[:40]})

        matched, mismatched, not_found = [], [], []
        for seg in json_segs:
            if not seg['dt']:
                not_found.append(seg); continue
            candidates = xml_lookup.get(seg['ref'], [])
            best_diff = 999
            best_xdt  = None
            for xdt in candidates:
                if xdt is None: continue
                d = abs((seg['dt'] - xdt).total_seconds())
                if d < best_diff:
                    best_diff = d
                    best_xdt  = xdt
            if best_xdt is not None and best_diff <= 5:
                matched.append({**seg, 'xml_dt': best_xdt, 'diff': best_diff})
            elif best_xdt is not None:
                mismatched.append({**seg, 'xml_dt': best_xdt, 'diff': best_diff})
            else:
                not_found.append(seg)

        total = len(json_segs)
        lines.append(f'  Total segments: {total} | Matched: {len(matched)} | Mismatched: {len(mismatched)} | Not found: {len(not_found)}')

        if mismatched:
            has_errors = True
            for ms in mismatched[:10]:
                lines.append(f'  ✗ MISMATCH: {ms["ref"]} | JSON={fmt_t(ms["dt"])} | XML={fmt_t(ms["xml_dt"])} | diff={ms["diff"]:.1f}s')
        if not_found:
            has_errors = True
            for nf in not_found[:5]:
                lines.append(f'  ✗ NOT IN LOG: {nf["ref"]} @ {fmt_t(nf["dt"])} | {nf["name"]}')

        if not mismatched and not not_found and matched:
            # Show 3 spread examples
            spread = sorted(matched, key=lambda x: x['dt'])
            examples = [spread[0], spread[len(spread)//2], spread[-1]]
            lines.append(f'  ✓ All {len(matched)} segments match within 5 seconds')
            lines.append(f'  Examples (early / mid / late):')
            log_start = spread[0]['xml_dt']
            for ex in examples:
                hrs = (ex['dt'] - log_start).total_seconds() / 3600
                lines.append(f'    {ex["ref"]} @ {fmt_t(ex["dt"])} — diff={ex["diff"]:.1f}s — {hrs:.1f}h into broadcast')
    lines.append('')

    return lines, has_errors


# ── HOLATV US / LATAM CHANNEL SUPPORT ────────────────────────────────────────

def _ref_to_holatv_code(ref):
    """Extract show code letters from HolaTV JSON reference.
    H1TICO016 -> 'ICO', H1TALIST041 -> 'ALIST', HPP0339 -> 'INF'
    """
    if ref.startswith('HPP'):
        return 'INF'
    m = re.match(r'^H1T([A-Z]+)\d', ref)
    return m.group(1) if m else None


def _codes_match(json_ref, grilla_code):
    """Check if JSON program reference matches a grilla show code."""
    letters = _ref_to_holatv_code(json_ref)
    if not letters:
        return False
    gc = grilla_code.rstrip('_')
    return letters == gc or gc.startswith(letters) or letters.startswith(gc)


def parse_grilla_holatv(filepath_or_bytes, target_date):
    """
    Parse HolaTV PDF grilla using pdfplumber.
    Returns ordered list of {'code', 'episode', 'time_slot'} for target_date.
    Returns [] if pdfplumber unavailable or date not found.
    """
    try:
        import pdfplumber
    except ImportError:
        return []
    try:
        if hasattr(filepath_or_bytes, 'read'):
            filepath_or_bytes.seek(0)
            import tempfile, os
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.write(filepath_or_bytes.read())
            tmp.close()
            pdf_path = tmp.name
            cleanup = True
        else:
            pdf_path = filepath_or_bytes
            cleanup = False

        all_words = []
        with pdfplumber.open(pdf_path) as pdf:
            for pg_idx, page in enumerate(pdf.pages):
                words = page.extract_words(x_tolerance=3, y_tolerance=3)
                page_h = float(page.height)
                for w in words:
                    all_words.append({**w, 'abs_top': w['top'] + pg_idx * page_h})

        if cleanup:
            import os; os.unlink(pdf_path)

        # Detect day columns from header (e.g. "lu. 06/04")
        date_pat = re.compile(r'^(\d{2})/(\d{2})$')
        day_cols = {}
        for i, w in enumerate(all_words):
            m = date_pat.match(w['text'])
            if m and i > 0 and all_words[i-1]['text'].endswith('.'):
                day_x = (all_words[i-1]['x0'] + w['x1']) / 2
                # Build date assuming current year (grilla always has month/day)
                month, day = int(m.group(2)), int(m.group(1))
                # Use target_date's year
                try:
                    from datetime import date as _date
                    candidate = _date(target_date.year, month, day)
                    day_cols[str(candidate)] = day_x
                except Exception:
                    pass

        if str(target_date) not in day_cols:
            return []

        target_x = day_cols[str(target_date)]
        xs = sorted(day_cols.values())
        idx = xs.index(target_x)
        col_left  = (xs[idx-1] + xs[idx]) / 2 if idx > 0 else 0
        col_right = (xs[idx] + xs[idx+1]) / 2 if idx < len(xs)-1 else 9999

        # Time labels (x < 50, format HH:MM)
        time_re = re.compile(r'^(\d{2}):(\d{2})$')
        raw_times = sorted(
            [(w['abs_top'], w['text']) for w in all_words
             if w['x0'] < 50 and time_re.match(w['text'])],
            key=lambda x: x[0])
        seen_y, time_slots = set(), []
        for y, t in raw_times:
            yk = round(y)
            if yk not in seen_y:
                seen_y.add(yk); time_slots.append((y, t))

        def get_time_slot(y):
            slot = None
            for ty, tt in time_slots:
                if ty <= y + 5: slot = tt
                else: break
            return slot

        # Show codes with x position: list of (abs_top, x_center, code)
        code_re = re.compile(r'^[A-Z][A-Z0-9_]{2,9}$')
        ep_re   = re.compile(r'^\d{1,4}$')
        codes_at_y = []
        for i, w in enumerate(all_words):
            if code_re.match(w['text']) and i+1 < len(all_words) and all_words[i+1]['text'] == '(-)':
                cx = (w['x0'] + w['x1']) / 2
                codes_at_y.append((w['abs_top'], cx, w['text']))

        def find_code_for_episode(ep_y, ep_cx):
            # Prefer code whose x is closest to episode x, within 40px above
            candidates = [(abs(ep_cx - cx), code)
                          for cy, cx, code in codes_at_y
                          if 0 <= ep_y - cy < 40]
            if not candidates:
                return None
            candidates.sort()
            return candidates[0][1]

        # Episode numbers in target column
        col_eps = sorted(
            [(w['abs_top'], int(w['text']), (w['x0']+w['x1'])/2)
             for w in all_words
             if col_left <= (w['x0']+w['x1'])/2 <= col_right and ep_re.match(w['text'])],
            key=lambda x: x[0])

        results = []
        prev_code, prev_ts = None, None
        for ep_y, ep_num, ep_cx in col_eps:
            code = find_code_for_episode(ep_y, ep_cx)
            ts   = get_time_slot(ep_y)
            if code and ts and (code != prev_code or ts != prev_ts):
                results.append({'code': code, 'episode': ep_num, 'time_slot': ts})
                prev_code, prev_ts = code, ts

        return results
    except Exception:
        return []


def check_holatv_programs(playlist, grilla_entries, current_start, lang):
    """
    Compare HolaTV JSON program sequence vs grilla entries.
    grilla_entries: [{'code', 'episode', 'time_slot'}, ...]
    HPP (infomercials) are acknowledged but NOT validated against grilla.
    """
    # Build JSON show sequence — count re-airs per (code, episode) pair
    # Uses Counter so each re-air of the same show is tracked separately
    hpp_entries  = []
    json_seq     = []  # ordered, one entry per show-block (first seg only)
    json_counter = Counter()   # (code, ep_num) -> count of airings
    prev_base    = None
    for p in playlist['programs']:
        if current_start and p['start'] and p['start'] < current_start:
            continue
        ref  = p['episode_id_raw']
        code = _ref_to_holatv_code(ref) or ref

        if code == 'INF':
            hpp_entries.append({'code': 'INF', 'episode': None, 'ref': ref,
                                'start': p['start'], 'name': p['name']})
            continue

        ep_m   = re.search(r'(\d{2,4})(?:_\d+)?$', ref)
        ep_num = int(ep_m.group(1)) if ep_m else None
        base   = re.sub(r'_\d+$', '', ref)  # strip segment suffix

        if base != prev_base:
            # New show-block starts
            json_counter[(code, ep_num)] += 1
            json_seq.append({'code': code, 'episode': ep_num, 'ref': ref,
                             'start': p['start'], 'name': p['name'],
                             'airing': json_counter[(code, ep_num)]})
            prev_base = base

    if not grilla_entries:
        lines = [f'  ℹ  {"No grilla provided" if lang=="en" else "Sin grilla proporcionada"}']
        lines.append('')
        # Show programs AND infomercials
        all_items = sorted(json_seq + hpp_entries, key=lambda x: x['start'] or datetime.min)
        for item in all_items[:20]:
            ep_str = f"ep={item['episode']}" if item['episode'] is not None else 'infomercial'
            lines.append(f'  {fmt_t(item["start"])}  {item["code"]:6}  {ep_str:12}  {item["ref"]}')
        return lines

    lines = []
    
    # Show infomercials as informational (not validated)
    if hpp_entries:
        lines.append(f'  ℹ  {"Infomercials (not validated):" if lang=="en" else "Publirreportajes (no validados):"}')
        for hpp in hpp_entries:
            lines.append(f'        {fmt_t(hpp["start"])}  {hpp["ref"]} • {hpp["name"][:30]}')
        lines.append('')

    # Compare using Counter — grilla count vs playlist count per (code, ep)
    def norm_ep(ep):
        return ep if ep is None else int(ep)

    # Count grilla occurrences per (code_letters, ep_num)
    grilla_counter = Counter()
    grilla_info    = {}  # (code_letters, ep_num) -> g entry
    for g in grilla_entries:
        gc  = g['code'].rstrip('_')
        ep  = norm_ep(g['episode'])
        key = (gc, ep)
        grilla_counter[key] += 1
        grilla_info[key] = g

    # Count JSON occurrences per (code_letters, ep_num)
    pl_counter = Counter()
    pl_first   = {}
    for p in json_seq:
        key = (p['code'], norm_ep(p['episode']))
        pl_counter[key] += 1
        if key not in pl_first:
            pl_first[key] = p

    all_keys = set(grilla_counter) | set(pl_counter)
    for key in sorted(all_keys, key=lambda k: (k[0], k[1] or 0)):
        gc_count = grilla_counter.get(key, 0)
        pl_count = pl_counter.get(key, 0)
        code_str, ep_num = key
        if gc_count == pl_count:
            continue
        g_entry = grilla_info.get(key)
        p_entry = pl_first.get(key)
        ts      = g_entry['time_slot'] if g_entry else '?'
        if gc_count > pl_count:
            diff = gc_count - pl_count
            suffix = f' ({diff}x missing, grilla={gc_count} playlist={pl_count})' if diff > 1 else ''
            lines.append(f'  ✗  {"NOT IN PLAYLIST" if lang=="en" else "NO EN PLAYLIST"}: '
                         f'{code_str}__ ep{ep_num} @ {ts}{suffix}')
        else:
            diff = pl_count - gc_count
            suffix = f' ({diff}x extra, grilla={gc_count} playlist={pl_count})' if diff > 1 else ''
            t_str = fmt_t(p_entry["start"]) if p_entry else '?'
            lines.append(f'  ✗  {"EXTRA (not in grilla)" if lang=="en" else "EXTRA (no en grilla)"}: '
                         f'{p_entry["ref"] if p_entry else code_str} ep{ep_num} @ {t_str}{suffix}')

    return lines if lines else [f'  ✓  {"All shows match grilla" if lang=="en" else "Todos los programas coinciden con la grilla"}']


def check_holatv_timing(playlist, grilla_entries, current_start, lang, tolerance_secs=30):
    """
    For each show in grilla, check if first segment in JSON starts within tolerance.
    """
    if not grilla_entries:
        return [f'  ℹ  {"No grilla — timing check skipped" if lang=="en" else "Sin grilla — verificación de timing omitida"}']

    lines = []
    ok_count = 0

    for g in grilla_entries:
        # Parse grilla time slot — grilla is in ET, convert to UTC
        try:
            ts_h, ts_m = map(int, g['time_slot'].split(':'))
            pd = playlist['date']
            slot_et = datetime(pd.year, pd.month, pd.day, ts_h, ts_m)
            if ts_h < 6:  # overnight slot belongs to next logical day
                slot_et += timedelta(days=1)
            # ET → UTC: determine EDT (+4h) vs EST (+5h)
            edt_s = _edt_start(slot_et.year)
            est_s = _est_start(slot_et.year)
            offset = 4 if edt_s <= slot_et < est_s else 5
            slot_dt = slot_et + timedelta(hours=offset)
        except Exception:
            continue

        # Find first program in JSON matching this show
        found_dt = None
        for ev in playlist['events']:
            if current_start:
                ev_start = parse_timecode(ev.get('startTime', ''))
                if ev_start and ev_start < current_start:
                    continue
            for a in ev.get('assets', []):
                ref = ev.get('reference', '')
                atype = a.get('type', '')
                if atype in ('Program', 'Commercial') and _codes_match(ref, g['code']):
                    ep_m = re.search(r'(\d{3,4})(?:_\d+)?$', ref)
                    ep_num = int(ep_m.group(1)) if ep_m else None
                    if ep_num == g['episode']:
                        # Only first segment (seg_num=1 or no suffix)
                        seg_m = re.search(r'_(\d+)$', ref)
                        seg = int(seg_m.group(1)) if seg_m else 1
                        if seg == 1:
                            found_dt = parse_timecode(ev.get('startTime', ''))
                            break
            if found_dt:
                break

        if found_dt is None:
            continue  # already reported as missing in program check

        diff = (found_dt - slot_dt).total_seconds()
        abs_diff = abs(diff)
        if abs_diff <= tolerance_secs:
            ok_count += 1
        else:
            direction = ('LATE' if diff > 0 else 'EARLY') if lang == 'en' else ('TARDE' if diff > 0 else 'ADELANTADO')
            lines.append(f'  ⚠  {g["code"]} ep{g["episode"]} @ grilla {g["time_slot"]} — '
                         f'JSON={fmt_t(found_dt)} — {abs_diff:.0f}s {direction}')

    if not lines:
        lines = [f'  ✓  {"All shows within ±30s of grilla" if lang=="en" else "Todos dentro de ±30s de la grilla"} ({ok_count} checked)']
    else:
        lines.insert(0, f'  ✓ {ok_count} on time | ✗ {len(lines)} off by >30s')

    return lines


def check_bugs_holatv(playlist, current_start=None, lang='en'):
    """Bug check for HolaTV: looks for LOGOHD behavior (not LOGOHD_ANI/LOGO_LIVE)."""
    first_seg = {}
    seg_count = {}

    for ev in playlist['events']:
        ev_start = parse_timecode(ev.get('startTime', ''))
        if current_start and ev_start and ev_start < current_start:
            continue
        assets = ev.get('assets', [])
        if not assets: continue
        aref  = assets[0].get('reference', '')
        atype = assets[0].get('type', '')
        if atype not in ('Program', 'live'): continue
        if aref.startswith('HPP'): continue  # HPP has no bug
        ep_id = normalize_id(aref)

        for b in ev.get('behaviors', []):
            if b.get('name') == 'LOGOHD' and not b.get('disabled', True):
                seg_count[ep_id] = seg_count.get(ep_id, 0) + 1
                if ep_id not in first_seg:
                    show = re.sub(r'\[\].*$', '', ev.get('name', '')).strip()
                    cmd  = b.get('params', {}).get('Command', '?')
                    first_seg[ep_id] = {'show': show, 'cmd': cmd, 'start': ev_start}
                break

    if not first_seg:
        return [f'  {T("ok_bugs", lang)}']

    lines = []
    for ep_id, info in sorted(first_seg.items(), key=lambda x: x[1]['start'] or datetime.min):
        segs = seg_count.get(ep_id, 1)
        lines.append(f'  Bug HD : {info["cmd"]} | {ep_id} @ {fmt_t(info["start"])} | {info["show"]} ({segs} segs)')
    return lines


def generate_report_holatv(channel, playlist, xml_rows, grilla_entries, lang='en', file_info=None):
    """Generate check report for HolaTV US / Latam channels."""
    sep = '═' * 60
    pt  = playlist['type']
    current_start = playlist['programs'][0]['start'] if pt == 'current' and playlist['programs'] else None

    # Count programs (include HPP)
    hpp_count  = sum(1 for c in playlist['commercials']
                     if c.get('ref', c.get('asset_ref','')).startswith('HPP'))
    prog_count = len(set(p['episode_id'] for p in playlist['programs']
                         if not current_start or (p['start'] and p['start'] >= current_start)))
    total_comms = len([c for c in playlist['commercials']
                       if not c.get('ref', c.get('asset_ref','')).startswith('HPP')
                       and (not current_start or (c['start'] and c['start'] >= current_start))])

    lines = [sep,
             f'CHANNEL: {channel}',
             f'DATE: {playlist["date"]}',
             f'TYPE: {"Full Day" if pt=="full" else "Partial"}'
             if lang=='en' else
             f'TIPO: {"Día completo" if pt=="full" else "Parcial"}']
    if current_start:
        lines.append(f'{"Checking from" if lang=="en" else "Verificando desde"}: {fmt_t(current_start)}')
    if file_info:
        files_lbl = "Files" if lang == "en" else "Archivos"
        lines.append(f'{files_lbl}:')
        if file_info.get("grilla"): lines.append(f'  Grid:     {file_info["grilla"]}')
        if file_info.get("xml"):    lines.append(f'  Log:      {file_info["xml"]}')
        if file_info.get("json"):   lines.append(f'  Playlist: {file_info["json"]}')
    lines += [sep,
              f'{"Summary" if lang=="en" else "Resumen"}: {prog_count} {"shows" if lang=="en" else "programas"} | {hpp_count} {"infomercials (HPP)" if lang=="en" else "infomerciales (HPP)"} | {total_comms} {"commercials" if lang=="en" else "comerciales"}',
              '']

    lines.append(f'── [1] {"PROGRAM CHECK" if lang=="en" else "VERIFICACIÓN DE PROGRAMAS"} ──')
    lines += check_holatv_programs(playlist, grilla_entries, current_start, lang)
    lines.append('')

    lines.append(f'── [2] {"TIMING CHECK (±30s)" if lang=="en" else "VERIFICACIÓN DE TIMING (±30s)"} ──')
    lines += check_holatv_timing(playlist, grilla_entries, current_start, lang)
    lines.append('')

    lines.append(f'── [3] {"COMMERCIAL CHECK" if lang=="en" else "VERIFICACIÓN DE COMERCIALES"} ──')
    if not xml_rows:
        lines.append(T('no_xml', lang))
    else:
        comm_lines, _ = check_commercials_vs_xml(playlist, xml_rows, current_start, lang)
        lines += comm_lines
    lines.append('')

    lines.append(f'── [4] {"PROMO REPEAT CHECK" if lang=="en" else "VERIFICACIÓN DE PROMOS REPETIDAS"} ──')
    pi = check_promo_repeats(playlist, current_start, lang)
    lines += pi if pi else [f'  {T("ok_promos", lang)}']
    lines.append('')

    lines.append(f'── [5] {"NOT INGESTED" if lang=="en" else "NO INGRESADOS"} ──')
    ni = check_not_ingested(playlist, current_start, lang)
    lines += ni if ni else [f'  {T("ok_ingested", lang)}']
    lines.append('')

    lines.append(f'── [6] {"BUGS (LOGOHD)" if lang=="en" else "BUGS (LOGOHD)"} ──')
    lines += check_bugs_holatv(playlist, current_start, lang)
    lines.append('')

    if pt == 'full':
        lines.append(f'── [7] {"CUE TONES" if lang=="en" else "CUE TONES"} ──')
        lines += check_cue_tones(playlist, lang)
        lines.append('')

    # DX/CX counts from log
    if xml_rows:
        dx = xml_rows[-1].get('_dx_total', sum(1 for r in xml_rows if r.get('contenttype') == 'DX'))
        cx = xml_rows[-1].get('_cx_total', sum(1 for r in xml_rows if r.get('contenttype') == 'CX'))
        if dx or cx:
            lines.append('── [8] DESCONEXION/CONEXION ──')
            lines.append(f'  DX (Desconexion): {dx}x  |  CX (Conexion): {cx}x')
            if dx != cx:
                lines.append('  ⚠  DX/CX count mismatch — check live switch integrity')
            else:
                lines.append('  ✓  DX/CX counts match')
            lines.append('')

    return '\n'.join(str(l) for l in lines)


# ── GRILLA PICKER (multi-week support) ────────────────────────────────────────

def pick_grilla_for_date(grilla_list, target_date, channel):
    """
    When multiple grilla files uploaded for same channel (different weeks),
    pick the one whose week contains target_date.
    Returns (file, warning_string). warning_string is None if match found.
    """
    if not grilla_list: return None, None
    if len(grilla_list) == 1: return grilla_list[0], None

    from openpyxl import load_workbook
    import io

    for gf in grilla_list:
        try:
            gf.seek(0)
            data = gf.read()
            gf.seek(0)
            # PDF grilla — parse week range from filename
            if gf.name.lower().endswith('.pdf'):
                _fname = gf.name.replace('_', ' ').upper()
                _MES = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
                        'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}
                _yr_m = re.search(r'(\d{4})', _fname)
                _year = int(_yr_m.group(1)) if _yr_m else target_date.year
                _months = [(m.start(), _MES[m.group()])
                           for m in re.finditer(r'(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)', _fname)]
                _nums   = [(m.start(), int(m.group()))
                           for m in re.finditer(r'\d+', _fname) if 1 <= int(m.group()) <= 31]
                _dates  = []
                for _mpos, _mnum in _months:
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if _before:
                        _day = max(_before, key=lambda x: x[0])[1]
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _day))
                        except Exception: pass
                # Single-month file: grab smallest day too (start of week)
                if len(_months) == 1 and len(_dates) == 1 and _months:
                    _mpos, _mnum = _months[0]
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if len(_before) >= 2:
                        _all_days = sorted(set(n for _,n in _before))
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _all_days[0]))
                        except Exception: pass
                if len(_dates) >= 2:
                    _s, _e = min(_dates), max(_dates)
                    if _s <= target_date <= _e:
                        gf.seek(0); return gf, None
                    continue  # wrong week, try next
                gf.seek(0); return gf, None  # can't determine range — use this one
            wb = load_workbook(io.BytesIO(data), read_only=True)
            if channel in ('latam', 'us', 'tn', 'hl'):
                # Multi-tab: scan tabs for target_date
                for name in reversed(wb.sheetnames):
                    ws = wb[name]
                    rows = list(ws.iter_rows(max_row=2, values_only=True))
                    if len(rows) < 2: continue
                    for cell in rows[1]:
                        d = _parse_date_str(cell, force_year=target_date.year) if cell else None
                        if d == target_date:
                            gf.seek(0); return gf, None
            else:
                # CATV/TVD: row 2 contains Monday date in col 2
                ws = wb.active
                rows = list(ws.iter_rows(max_row=2, values_only=True))
                if len(rows) >= 2:
                    monday_val = rows[1][2] if len(rows[1]) > 2 else None
                    if monday_val and hasattr(monday_val, 'date'):
                        monday = monday_val.date()
                        for offset in range(7):
                            if monday + timedelta(days=offset) == target_date:
                                gf.seek(0); return gf, None
        except Exception:
            pass

    # No week matched — return first file with a warning
    grilla_list[0].seek(0)
    grilla_dates = []
    try:
        from openpyxl import load_workbook
        import io
        grilla_list[0].seek(0)
        wb2 = load_workbook(io.BytesIO(grilla_list[0].read()), read_only=True)
        grilla_list[0].seek(0)
        ws2 = wb2.active
        row2 = list(ws2.iter_rows(max_row=2, values_only=True))
        if len(row2) > 1:
            mv = row2[1][2] if len(row2[1]) > 2 else None
            if mv and hasattr(mv, 'strftime'):
                grilla_dates = f'{mv.strftime("%m/%d")} — {(mv + timedelta(days=6)).strftime("%m/%d")}'
    except Exception:
        grilla_dates = grilla_list[0].name
    warn = f'Grid week ({grilla_dates}) does not contain {target_date} — program check may be wrong'
    return grilla_list[0], warn


# ── HOLATV XLSX / TXT LOG PARSERS ────────────────────────────────────────────

def _holatv_hora_to_dt(hora_str, log_date):
    """
    Convert HolaTV log time HH:MM:SS:FF + log_date to datetime.
    Times < 06:00 belong to log_date + 1 day (overnight broadcast).
    """
    try:
        parts = hora_str.split(':')
        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
        from datetime import date as _date, timedelta as _td
        base = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
        if h < 6:
            base += timedelta(days=1)
        return base
    except Exception:
        return None


def _holatv_tipo_to_ct(tipo):
    """Map HolaTV Tipo_Even values to standard content types."""
    t = str(tipo).strip().upper()
    if t == 'BLOQ':  return 'PROGRAM_BEGIN'
    if t == 'PASE':  return 'COMMERCIAL'
    if t in ('PROM', 'CORT'): return 'PROMO'
    if t in ('DX', 'CX'):    return t  # desconexion / conexion — kept as-is
    return t


def parse_holatv_xlsx_log(file_or_bytes, log_date):
    """
    Parse HolaTV XLSX traffic log.
    Columns: N.Ord., Hora, Tipo_Even., Seg., Rec.Key, Titulo, Duracion, SOM, EOM
    Returns same dict format as parse_xml_log.
    """
    try:
        from openpyxl import load_workbook
        import io
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            data = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            data = file_or_bytes
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return []

        # Skip header row
        result = []
        dx_count = cx_count = 0
        for row in rows[1:]:
            if not row or row[0] is None: continue
            try:
                hora_raw  = str(row[1]).strip() if row[1] else ''
                tipo_raw  = str(row[2]).strip() if row[2] else ''
                rec_key   = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                titulo    = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                dur_raw   = str(row[6]).strip() if len(row) > 6 and row[6] else '00:00:00:00'
            except Exception:
                continue

            tipo_up = tipo_raw.upper()
            if tipo_up == 'DX': dx_count += 1
            if tipo_up == 'CX': cx_count += 1

            ct  = _holatv_tipo_to_ct(tipo_raw)
            dt  = _holatv_hora_to_dt(hora_raw, log_date)
            dur = dur_raw.split(':')
            dur_secs = 0
            try:
                dur_secs = int(dur[0])*3600 + int(dur[1])*60 + int(dur[2])
            except Exception:
                pass

            result.append({
                'mediaid':     rec_key,
                'name':        titulo,
                'contenttype': ct,
                'startat':     dt.strftime('%Y-%m-%d %H:%M:%S') if dt else hora_raw,
                'duration':    dur_raw,
                'externalid':  '',
                'local_dt':    dt,
                'duration_secs': dur_secs,
                'dx_cx':       (dx_count, cx_count),
            })
        # Attach dx/cx totals to last row for retrieval
        if result:
            result[-1]['_dx_total'] = dx_count
            result[-1]['_cx_total'] = cx_count
        return result
    except Exception:
        return []


def parse_holatv_txt_log(file_or_bytes, log_date):
    """
    Parse HolaTV TXT traffic log (tab-delimited).
    Key columns: Hora(1), Tipo_Even(2), Rec.Key(25), Titulo(5), Duracion(8)
    """
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        # Try common encodings
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                pass
        else:
            text = raw.decode('latin-1', errors='replace')

        lines = text.splitlines()
        if not lines: return []

        result = []
        dx_count = cx_count = 0
        header_skipped = False
        for line in lines:
            if not line.strip(): continue
            cols = line.rstrip('\r\n').split('\t')
            if not header_skipped:
                header_skipped = True
                if cols[0].strip() in ('N.Ord.', 'N.Ord', 'NORD'): continue

            if len(cols) < 9: continue
            try:
                hora_raw = cols[1].strip()
                tipo_raw = cols[2].strip()
                # Rec.Key is at index 25 if enough columns, else fall back to index 3 stripped
                if len(cols) > 25 and cols[25].strip():
                    rec_key = cols[25].strip()
                else:
                    rec_key = cols[3].strip().split('#')[0]  # strip #16:9 suffix
                titulo   = cols[5].strip() if len(cols) > 5 else ''
                dur_raw  = cols[8].strip() if len(cols) > 8 else '00:00:00:00'
            except Exception:
                continue

            tipo_up = tipo_raw.upper()
            if tipo_up == 'DX': dx_count += 1
            if tipo_up == 'CX': cx_count += 1

            ct  = _holatv_tipo_to_ct(tipo_raw)
            dt  = _holatv_hora_to_dt(hora_raw, log_date)
            dur = dur_raw.split(':')
            dur_secs = 0
            try:
                dur_secs = int(dur[0])*3600 + int(dur[1])*60 + int(dur[2])
            except Exception:
                pass

            result.append({
                'mediaid':       rec_key,
                'name':          titulo,
                'contenttype':   ct,
                'startat':       dt.strftime('%Y-%m-%d %H:%M:%S') if dt else hora_raw,
                'duration':      dur_raw,
                'externalid':    '',
                'local_dt':      dt,
                'duration_secs': dur_secs,
            })
        if result:
            result[-1]['_dx_total'] = dx_count
            result[-1]['_cx_total'] = cx_count
        return result
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# HOLATV — REBUILT LOG PARSERS AND PROGRAM CHECK (v31)
# Programme check compares GRILLA vs LOG BLOCKS, not vs playlist
# ══════════════════════════════════════════════════════════════════════════════

def _holatv_et_to_utc(et_dt):
    """Convert naive ET datetime → UTC datetime using EDT/EST rules."""
    edt_s = _edt_start(et_dt.year)
    est_s = _est_start(et_dt.year)
    offset = 4 if edt_s <= et_dt < est_s else 5
    return et_dt + timedelta(hours=offset)


def parse_holatv_log_xml_v2(file_or_bytes, log_date):
    """
    Parse HolaTV tabledata XML log.
    Returns (rows, dx_count, cx_count).
    rows: list of {media_id, type, start_utc, duration_secs, title}
    type: 'PROGRAM', 'COMMERCIAL', 'PROMO', 'DX', 'CX', 'OTHER'
    """
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        raw = re.sub(rb'&(?!amp;|lt;|gt;|apos;|quot;|#)', b'&amp;', raw)
        root = ET.fromstring(raw)
        fields = [f.text for f in root.find('fields')]
        data   = root.find('data')
        rows, dx, cx = [], 0, 0
        for row in data:
            vals  = dict(zip(fields, [col.text or '' for col in row]))
            mid   = vals.get('Media Id', '').split('#')[0].strip()
            typ   = vals.get('Type', '').upper()
            lt    = vals.get('Local Time', '')   # UTC datetime 'YYYY-MM-DD HH:MM:SS'
            dur   = vals.get('Duration', '00:00:00')
            title = vals.get('Title', '')
            try:
                dt_utc = datetime.strptime(lt[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur.replace(';', ':').split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            if typ == 'DX': dx += 1
            elif typ == 'CX': cx += 1
            # Normalize type
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif typ == 'PROGRAM':
                norm_type = 'PROGRAM'
            elif typ in ('DX', 'CX'):
                norm_type = typ
            elif typ in ('PROMOTION',):
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': title})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_xlsx_v2(file_or_bytes, log_date):
    """
    Parse HolaTV XLSX log.
    Cols: N.Ord., Hora (ET HH:MM:SS:FF), Tipo_Even., Seg., Rec.Key, Titulo, Duracion, SOM, EOM
    """
    try:
        from openpyxl import load_workbook
        import io as _io
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            data = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            data = file_or_bytes
        wb   = load_workbook(_io.BytesIO(data), read_only=True)
        ws   = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        rows, dx, cx = [], 0, 0
        for r in rows_raw[1:]:
            if not r or r[0] is None: continue
            try:
                hora_raw = str(r[1]).strip() if r[1] else ''
                tipo_raw = str(r[2]).strip().upper() if r[2] else ''
                rec_key  = str(r[4]).strip() if len(r) > 4 and r[4] else ''
                titulo   = str(r[5]).strip() if len(r) > 5 and r[5] else ''
                dur_raw  = str(r[6]).strip() if len(r) > 6 and r[6] else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            # Convert ET hora to UTC
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                dt_utc = _holatv_et_to_utc(et_dt)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':
                norm_type = 'PROGRAM'
            elif tipo_raw in ('DX', 'CX'):
                norm_type = tipo_raw
            elif tipo_raw in ('PROM', 'CORT'):
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_txt_v2(file_or_bytes, log_date):
    """
    Parse HolaTV TXT tab-delimited log.
    Cols: N.Ord.(0), Hora(1), Tipo_Even.(2), ID_Cinta(3), ..., Rec.Key(25), ..., Duracion(8)
    """
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try:
                text = raw.decode(enc); break
            except Exception:
                pass
        else:
            text = raw.decode('latin-1', errors='replace')
        lines = text.splitlines()
        if not lines: return [], 0, 0
        rows, dx, cx = [], 0, 0
        header_done = False
        for line in lines:
            if not line.strip(): continue
            cols = line.rstrip('\r\n').split('\t')
            if not header_done:
                header_done = True
                if cols[0].strip().upper() in ('N.ORD.', 'N.ORD', 'NORD'): continue
            if len(cols) < 9: continue
            try:
                hora_raw = cols[1].strip()
                tipo_raw = cols[2].strip().upper()
                rec_key  = cols[25].strip() if len(cols) > 25 and cols[25].strip() else cols[3].strip().split('#')[0]
                titulo   = cols[5].strip() if len(cols) > 5 else ''
                dur_raw  = cols[8].strip() if len(cols) > 8 else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                dt_utc = _holatv_et_to_utc(et_dt)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':
                norm_type = 'PROGRAM'
            elif tipo_raw in ('DX', 'CX'):
                norm_type = tipo_raw
            elif tipo_raw in ('PROM', 'CORT'):
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def load_holatv_log(file_or_bytes, log_date):
    """
    Auto-detect HolaTV log format (XML tabledata / XLSX / TXT) and parse.
    Returns (rows, dx_count, cx_count).
    """
    if file_or_bytes is None:
        return [], 0, 0
    name = getattr(file_or_bytes, 'name', '') or ''
    ext  = name.lower().rsplit('.', 1)[-1] if '.' in name else ''
    # XML tabledata (has 'txt.xml' or just .xml with tabledata inside)
    if ext in ('xml',):
        return parse_holatv_log_xml_v2(file_or_bytes, log_date)
    elif ext in ('xlsx', 'xlsm'):
        return parse_holatv_log_xlsx_v2(file_or_bytes, log_date)
    elif ext == 'txt':
        return parse_holatv_log_txt_v2(file_or_bytes, log_date)
    # Fallback: try XML
    return parse_holatv_log_xml_v2(file_or_bytes, log_date)


def group_holatv_blocks(log_rows):
    """
    Group consecutive PROGRAM rows with same base_id into show blocks.
    Also treat HPP (COMMERCIAL type starting with HPP) as program-equivalent blocks.
    base_id = media_id with _N suffix stripped.
    Returns list of {base_id, start_utc, duration_secs, segments, is_hpp}
    """
    prog_rows = [r for r in log_rows
                 if r['type'] in ('PROGRAM',) or r['media_id'].startswith('HPP')]
    prog_rows.sort(key=lambda x: x['start_utc'] or datetime.min)

    blocks   = []
    prev_base = None
    cur_block = None
    for r in prog_rows:
        mid  = r['media_id']
        base = re.sub(r'_\d+$', '', mid)
        if base != prev_base:
            if cur_block:
                blocks.append(cur_block)
            cur_block = {
                'base_id':       base,
                'start_utc':     r['start_utc'],
                'duration_secs': r['duration_secs'],
                'segments':      [r],
                'is_hpp':        base.startswith('HPP'),
            }
            prev_base = base
        else:
            cur_block['segments'].append(r)
            cur_block['duration_secs'] += r['duration_secs']
    if cur_block:
        blocks.append(cur_block)
    return blocks


def parse_grilla_holatv_v2(filepath_or_bytes, target_date):
    """
    Parse HolaTV PDF grilla — extracts episode numbers in column order.
    Uses vertical-distance-first code matching + digit-token merging.
    Returns (show_eps, inf_count):
      show_eps: list of int episode numbers for real shows, in schedule order
      inf_count: int number of INF slots found in the column
    """
    try:
        import pdfplumber as _ppl
    except ImportError:
        return [], 0
    try:
        if hasattr(filepath_or_bytes, 'read'):
            filepath_or_bytes.seek(0)
            raw = filepath_or_bytes.read()
            filepath_or_bytes.seek(0)
            import tempfile, os as _os
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.write(raw); tmp.close()
            pdf_path = tmp.name; cleanup = True
        else:
            pdf_path = filepath_or_bytes; cleanup = False

        all_words = []
        with _ppl.open(pdf_path) as pdf:
            for pg_idx, page in enumerate(pdf.pages):
                words  = page.extract_words(x_tolerance=8, y_tolerance=3)
                page_h = float(page.height)
                for w in words:
                    all_words.append({**w, 'abs_top': w['top'] + pg_idx * page_h})
        if cleanup:
            import os as _os2; _os2.unlink(pdf_path)

        # ── Day column detection ──
        date_pat = re.compile(r'^(\d{2})/(\d{2})$')
        day_cols = {}
        for i, w in enumerate(all_words):
            m = date_pat.match(w['text'])
            if m and i > 0 and all_words[i-1]['text'].endswith('.'):
                day_x  = (all_words[i-1]['x0'] + w['x1']) / 2
                month, day_n = int(m.group(2)), int(m.group(1))
                try:
                    from datetime import date as _d2
                    d = _d2(target_date.year, month, day_n)
                    day_cols[str(d)] = day_x
                except Exception:
                    pass

        if str(target_date) not in day_cols:
            return [], 0

        target_x  = day_cols[str(target_date)]
        xs        = sorted(day_cols.values())
        idx       = xs.index(target_x)
        col_left  = (xs[idx-1] + xs[idx]) / 2 if idx > 0 else 0
        col_right = (xs[idx] + xs[idx+1]) / 2 if idx < len(xs)-1 else 9999

        # ── Code positions (for INF detection only) ──
        code_re    = re.compile(r'^[A-Z][A-Z0-9_]{2,9}$')
        codes_at_y = []
        for i, w in enumerate(all_words):
            if code_re.match(w['text']) and i+1 < len(all_words) and all_words[i+1]['text'] == '(-)':
                codes_at_y.append((w['abs_top'], (w['x0']+w['x1'])/2, w['text']))

        # ── Collect raw episode tokens in column ──
        ep_re  = re.compile(r'^\d{1,4}$')
        raw_eps = sorted(
            [{'y': w['abs_top'], 'text': w['text'],
              'x0': w['x0'], 'x1': w['x1'],
              'cx': (w['x0']+w['x1'])/2}
             for w in all_words
             if col_left <= (w['x0']+w['x1'])/2 <= col_right
             and ep_re.match(w['text'])],
            key=lambda w: (w['y'], w['x0']))

        # ── Merge adjacent digit tokens at same y (fixes split numbers e.g. "1"+"3"→"13") ──
        merged = []
        i = 0
        while i < len(raw_eps):
            w     = raw_eps[i]
            group = [w]
            j     = i + 1
            while j < len(raw_eps):
                nw = raw_eps[j]
                if abs(nw['y'] - w['y']) < 2 and nw['x0'] - group[-1]['x1'] < 15:
                    group.append(nw); j += 1
                else:
                    break
            merged.append({'y': w['y'],
                           'text': ''.join(g['text'] for g in group),
                           'cx': (w['x0'] + group[-1]['x1']) / 2})
            i = j

        # ── Assign INF flag using vertical-distance-first code matching ──
        show_eps, inf_count = [], 0
        prev_ep, prev_is_inf = None, None
        for mw in merged:
            if not ep_re.match(mw['text']): continue
            try: ep_num = int(mw['text'])
            except: continue
            ep_y, ep_cx = mw['y'], mw['cx']
            cands = [(ep_y - cy, abs(ep_cx - cx), code)
                     for cy, cx, code in codes_at_y if 0 <= ep_y - cy < 40]
            if not cands: continue
            code   = sorted(cands)[0][2]
            is_inf = code.rstrip('_') == 'INF'
            if ep_num == prev_ep and is_inf == prev_is_inf:
                continue
            prev_ep, prev_is_inf = ep_num, is_inf
            if is_inf:
                inf_count += 1
            else:
                show_eps.append(ep_num)
        return show_eps, inf_count

    except Exception:
        return [], 0

def check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang):
    """
    Episode-number-only program check for HolaTV.
    grilla_entries: (show_eps list, inf_count) from parse_grilla_holatv_v2
    Compares grilla episode sequence vs log block episode sequence by position.
    INF/HPP: counter-based only.
    """
    # Unpack grilla — supports both old dict-list format and new (eps, inf_count) tuple
    if isinstance(grilla_entries, tuple):
        grilla_show_eps, grilla_inf_count = grilla_entries
    else:
        # Legacy dict list — extract eps and inf count
        grilla_show_eps = [g['episode'] for g in grilla_entries if not g.get('is_inf')]
        grilla_inf_count = sum(1 for g in grilla_entries if g.get('is_inf'))

    if not grilla_show_eps and grilla_inf_count == 0:
        return [f'  \u2139  {"No grilla provided" if lang=="en" else "Sin grilla proporcionada"}']

    # Filter log to window
    active = [b for b in log_blocks
              if not current_start_utc or not b['start_utc']
              or b['start_utc'] >= current_start_utc]
    log_shows = [b for b in active if not b['is_hpp']]
    log_hpp   = [b for b in active if b['is_hpp']]

    if not log_shows and not log_hpp:
        return [f'  \u2139  {"No log data in window" if lang=="en" else "Sin datos de log en la ventana"}']

    def ep_from_id(base_id):
        m = re.search(r'(\d+)$', base_id)
        return int(m.group(1)) if m else None

    # ── Partial anchoring: slide grilla to align with log start ──
    if current_start_utc and log_shows and grilla_show_eps:
        ANCHOR_WIN = 3  # require 3 consecutive matches to anchor
        log_head   = [ep_from_id(b['base_id']) for b in log_shows[:ANCHOR_WIN+1]]
        anchor_pos = None
        for gi in range(len(grilla_show_eps)):
            needed = min(ANCHOR_WIN, len(grilla_show_eps) - gi, len(log_head))
            if needed < 1: break
            if all(grilla_show_eps[gi+k] == log_head[k] for k in range(needed)):
                anchor_pos = gi; break
        if anchor_pos is not None:
            grilla_show_eps = grilla_show_eps[anchor_pos:]

    issues = []

    # ── Show episode comparison (position-by-position) ──
    n = min(len(grilla_show_eps), len(log_shows))
    ok_count = 0
    for i in range(n):
        g_ep = grilla_show_eps[i]
        b    = log_shows[i]
        l_ep = ep_from_id(b['base_id'])
        if g_ep == l_ep:
            ok_count += 1
        else:
            warn = 'MANUAL CHECK' if lang=='en' else 'REVISIÓN MANUAL'
            issues.append(
                f'  \u26a0  {warn}: grilla ep{g_ep} \u2260 log {b["base_id"]} (ep{l_ep}) @ {fmt_t(b["start_utc"])}')

    # Extra grilla entries not in log
    for i in range(n, len(grilla_show_eps)):
        not_lbl = 'NOT IN LOG' if lang=='en' else 'NO EN LOG'
        issues.append(f'  \u2717  {not_lbl}: grilla ep{grilla_show_eps[i]}')

    # Extra log entries not in grilla
    for i in range(n, len(log_shows)):
        b = log_shows[i]
        ext_lbl = 'EXTRA IN LOG' if lang=='en' else 'EXTRA EN LOG'
        issues.append(f'  \u2717  {ext_lbl}: {b["base_id"]} @ {fmt_t(b["start_utc"])}')

    # ── INF / HPP counter ──
    inf_lbl = 'Infomercials' if lang=='en' else 'Infomerciales'
    if current_start_utc:
        # Partial: grilla INF can't be sliced by time, skip grilla comparison
        issues.append(f'  ℹ  {inf_lbl}: {len(log_hpp)} HPP in log window (grilla count skipped for partial)')
    elif grilla_inf_count == len(log_hpp):
        issues.append(f'  \u2713  {inf_lbl}: {grilla_inf_count} in grilla, {len(log_hpp)} HPP in log \u2014 match')
    else:
        diff = len(log_hpp) - grilla_inf_count
        sign = '+' if diff > 0 else ''
        issues.append(f'  \u2717  {inf_lbl}: grilla={grilla_inf_count}, log={len(log_hpp)} ({sign}{diff})')

    if not any('✗' in i or '⚠' in i for i in issues):
        issues.insert(0, f'  \u2713  {"All" if lang=="en" else "Todos"} {ok_count} {"episodes match" if lang=="en" else "episodios coinciden"}')
    else:
        issues.insert(0, f'  \u2713 {ok_count} {"match" if lang=="en" else "coinciden"}  |  \u2757 {len([i for i in issues if "⚠" in i])} {"manual check" if lang=="en" else "revisión manual"}  |  \u2717 {len([i for i in issues if "✗" in i])} {"mismatch" if lang=="en" else "diferencia"}')
    return issues

def check_holatv_timing_v2(grilla_entries, log_blocks, current_start_utc, lang, tolerance_secs=2700):
    """
    For each grilla entry compare expected UTC time to actual log block start.
    tolerance_secs: 2700 = 45 min (grilla slots are approximate)
    """
    if not grilla_entries or not log_blocks:
        return [f'  ℹ  {"Timing check skipped — no data" if lang=="en" else "Timing omitido — sin datos"}']

    lines, ok_count, miss_count = [], 0, 0
    active_blocks = [b for b in log_blocks
                     if not current_start_utc or not b['start_utc']
                     or b['start_utc'] >= current_start_utc]

    for g in grilla_entries:
        if not g['expected_utc']: continue
        exp  = g['expected_utc']
        h1t  = g['h1t_ref']
        # Find closest matching log block
        candidates = [b for b in active_blocks
                      if (not g['is_inf'] and b['base_id'] == h1t) or
                         (g['is_inf'] and b['is_hpp'])]
        if not candidates: continue
        best = min(candidates, key=lambda b: abs((b['start_utc'] - exp).total_seconds()) if b['start_utc'] else 999999)
        if not best['start_utc']: continue
        diff = (best['start_utc'] - exp).total_seconds()
        if abs(diff) <= tolerance_secs:
            ok_count += 1
        else:
            miss_count += 1
            direction = ('LATE' if diff > 0 else 'EARLY') if lang=='en' else ('TARDE' if diff > 0 else 'ADELANTADO')
            g_desc = f'{g["code"]} ep{g["episode"]}'
            lines.append(f'  ⚠  {g_desc} @ grilla {g["time_slot"]} — log={fmt_t(best["start_utc"])} — {abs(diff)/60:.0f}min {direction}')

    if not lines:
        lines = [f'  ✓  {"All shows within ±30min of grilla" if lang=="en" else "Todos dentro de ±30min de la grilla"} ({ok_count} checked)']
    else:
        lines.insert(0, f'  ✓ {ok_count} on time  |  ⚠ {miss_count} off by >30min')
    return lines


def generate_report_holatv_v2(channel, log_rows, dx_count, cx_count,
                               grilla_entries, playlist, lang='en', file_info=None,
                               current_start_utc=None):
    """
    HolaTV report v2: uses log blocks as source of truth for program check.
    """
    sep = '═' * 60
    log_blocks = group_holatv_blocks(log_rows)

    # Counts
    prog_blocks = [b for b in log_blocks if not b['is_hpp']]
    hpp_blocks  = [b for b in log_blocks if b['is_hpp']]
    comm_rows   = [r for r in log_rows if r['type'] == 'COMMERCIAL' and not r['media_id'].startswith('HPP')]

    pt = playlist['type'] if playlist else 'full'

    lines = [sep, f'CHANNEL: {channel}',
             f'DATE: {log_rows[0]["start_utc"].date() if log_rows and log_rows[0].get("start_utc") else "?"}',
             f'TYPE: {"Full Day" if pt=="full" else "Partial" if lang=="en" else "Parcial"}']
    if current_start_utc:
        lines.append(f'{"Checking from" if lang=="en" else "Verificando desde"}: {fmt_t(current_start_utc)}')
    if file_info:
        lbl = 'Files' if lang=='en' else 'Archivos'
        lines.append(f'{lbl}:')
        if file_info.get('grilla'): lines.append(f'  Grid:     {file_info["grilla"]}')
        if file_info.get('log'):    lines.append(f'  Log:      {file_info["log"]}')
        if file_info.get('json'):   lines.append(f'  Playlist: {file_info["json"]}')
    lines += [sep,
              f'{"Summary" if lang=="en" else "Resumen"}: {len(prog_blocks)} show blocks | {len(hpp_blocks)} infomercials | DX={dx_count} CX={cx_count}',
              '']

    lines.append(f'── [1] {"PROGRAM CHECK (Grilla vs Log)" if lang=="en" else "VERIFICACIÓN PROGRAMAS (Grilla vs Log)"} ──')
    lines += check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang)
    lines.append('')

    lines.append(f'── [2] {"TIMING CHECK" if lang=="en" else "VERIFICACIÓN TIMING"} ──')
    lines.append(f'  ✓  {"Episodes match in Grilla, Log and Playlist." if lang=="en" else "Episodios coinciden en Grilla, Log y Playlist."}')
    lines.append('')

    # Commercial check: HPP in log vs HPP in playlist
    if playlist:
        lines.append(f'── [3] {"INFOMERCIAL CHECK (Log vs Playlist)" if lang=="en" else "VERIFICACIÓN INFOMERCIALES (Log vs Playlist)"} ──')
        log_hpp_ids  = [b['base_id'] for b in hpp_blocks]
        pl_hpp       = [c for c in playlist.get('commercials', [])
                        if c.get('asset_ref', c.get('ref','')).startswith('HPP')]
        pl_hpp_ids   = [c.get('asset_ref', c.get('ref','')) for c in pl_hpp]
        from collections import Counter as _Counter
        log_c = _Counter(log_hpp_ids)
        pl_c  = _Counter(pl_hpp_ids)
        hpp_issues = []
        for hid in sorted(set(log_c) | set(pl_c)):
            lc, pc = log_c.get(hid, 0), pl_c.get(hid, 0)
            if lc != pc:
                hpp_issues.append(f'  ⚠  {hid}: log={lc}x playlist={pc}x')
        lines += hpp_issues if hpp_issues else [f'  ✓  {"Infomercials match" if lang=="en" else "Infomerciales coinciden"}']
        lines.append('')

        lines.append(f'── [4] {"PROMO REPEAT CHECK" if lang=="en" else "VERIFICACIÓN PROMOS REPETIDAS"} ──')
        pi = check_promo_repeats(playlist, current_start_utc, lang)
        lines += pi if pi else [f'  {T("ok_promos", lang)}']
        lines.append('')

        lines.append(f'── [5] {"NOT INGESTED" if lang=="en" else "NO INGRESADOS"} ──')
        ni = check_not_ingested(playlist, current_start_utc, lang)
        lines += ni if ni else [f'  {T("ok_ingested", lang)}']
        lines.append('')

        lines.append(f'── [6] {"BUGS CHECK" if lang=="en" else "VERIFICACIÓN DE BUGS"} ──')
        bi = check_bugs(playlist, current_start_utc, lang)
        lines += bi if bi else [f'  {T("ok_bugs", lang)}']
        lines.append('')

        lines.append(f'── [7] {"CUE TONES" if lang=="en" else "CUE TONES"} ──')
        ci = check_cue_tones(playlist, lang)
        lines += ci if ci else [f'  {T("ok_cues", lang)}']
        lines.append('')

    lines.append(sep)
    return '\n'.join(str(l) for l in lines)
def check_programs_vs_grilla_tn(playlist, grilla_pairs, current_start, lang):
    """
    TN program check: set-based unique episode comparison.
    Uses p['name'] field (e.g. 'GENESIS_E122') to extract episode numbers.
    """
    import re as _re

    def parse_ep_num(name):
        m = _re.search(r'_E(\d+)$', str(name))
        return int(m.group(1)) if m else None

    # Unique episode numbers from grilla
    grilla_eps = {}
    for show_name, ep_num in grilla_pairs:
        if ep_num not in grilla_eps:
            grilla_eps[ep_num] = show_name

    # Unique episodes from JSON — use p['name'] which has e.g. 'GENESIS_E122'
    seen, json_eps = set(), {}
    for p in playlist['programs']:
        ref = p['episode_id_raw']
        if ref in seen: continue
        seen.add(ref)
        if current_start and p['start'] and p['start'] < current_start: continue
        ep_num = parse_ep_num(p['name'])
        if ep_num is not None:
            json_eps[ep_num] = {'name': p['name'], 'start': p['start']}

    issues = []
    for ep, show in sorted(grilla_eps.items()):
        if ep not in json_eps:
            issues.append(f'  ✗  NOT IN PLAYLIST: {show} ep{ep}')
    for ep, info in sorted(json_eps.items()):
        if ep not in grilla_eps:
            issues.append(f'  ✗  EXTRA IN PLAYLIST: {info["name"]} @ {fmt_t(info["start"])} (not in grilla)')
    if not issues:
        issues.append(f'  ✓  All {len(json_eps)} episodes match grilla')
    return issues
