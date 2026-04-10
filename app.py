"""
Broadcast Playlist Checker — Streamlit App v6
"""
import streamlit as st
import json
from datetime import datetime
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from checker import (
    parse_json_playlist, parse_xml_log, parse_xml_log_tn, parse_grilla,
    generate_report, check_promo_repeats, detect_files,
    parse_sony_xml_log, check_sony, pair_sony_files, SONY_CHANNEL_MAP,
    parse_sony_json_markers,
    load_holatv_log, group_holatv_blocks, parse_grilla_holatv_v2,
    generate_report_holatv_v2, pick_grilla_for_date,
)

APP_VERSION = "v30.5"

st.set_page_config(page_title='Broadcast Playlist Checker', layout='wide')

SONY_EMOJI = {
    'A1':'🅰️','A2':'🅰️','A3':'🅰️','A4':'🅰️','A5':'🅰️','A6':'🅰️',
    'F1':'🎬','F4':'🎬',
    'S1':'📡','S2':'📡','S3':'📡','S4':'📡','S5':'📡','S6':'📡',
}
SONY_L = {
    'markers_hdr':  {'en': '── [1] MARKERS ──',                    'es': '── [1] MARCADORES ──'},
    'no_marker':    {'en': '  ℹ  No markers found (partial/current playlist)', 'es': '  ℹ  Sin marcadores (playlist parcial/actual)'},
    'log_hdr':      {'en': '── [2] LOG FILE MATCH ──',             'es': '── [2] VERIFICACIÓN DE ARCHIVO LOG ──'},
    'ep_hdr':       {'en': '── [3] ENDPOINT CHECK ──',             'es': '── [3] VERIFICACIÓN DE PUNTO FINAL ──'},
    'seg_hdr':      {'en': '── [4] SEGMENT TIMING CHECK (≤5s tolerance) ──', 'es': '── [4] VERIFICACIÓN DE TIEMPOS (tolerancia ≤5s) ──'},
    'pl_full':      {'en': 'FULL (marker present)',                 'es': 'COMPLETO (marcador presente)'},
    'pl_partial':   {'en': 'CURRENT (partial)',                     'es': 'ACTUAL (parcial)'},
    'pl_none':      {'en': '— no JSON —',                          'es': '— sin JSON —'},
    'no_json':      {'en': '  ℹ  Log provided but no matching JSON found', 'es': '  ℹ  Log provisto pero sin JSON correspondiente'},
    'no_log':       {'en': '  ℹ  JSON found but no matching log provided', 'es': '  ℹ  JSON encontrado pero sin log correspondiente'},
}
def SL(key, lang): return SONY_L.get(key, {}).get(lang, SONY_L.get(key, {}).get('en', key))

# ── LANGUAGE ──────────────────────────────────────────────────────────────────
lang = st.radio('🌐', ['English', 'Español'], horizontal=True, label_visibility='collapsed')
lang = 'es' if lang == 'Español' else 'en'

L = {
    'title':    {'en': '📋 Broadcast Playlist Checker',             'es': '📋 Verificador de Playlist'},
    'upload':   {'en': 'Drop all files here — auto-detected (JSON, XML, XLSX)', 'es': 'Arrastra archivos aquí — detección automática (JSON, XML, XLSX)'},
    'run':      {'en': '▶  Run Check',                              'es': '▶  Verificar'},
    'dl':       {'en': '⬇ Download Report (.txt)',                  'es': '⬇ Descargar Reporte (.txt)'},
    'detected': {'en': '**Detected files:**',                       'es': '**Archivos detectados:**'},
    'unknown':  {'en': '⚠ Unrecognized:',                          'es': '⚠ No reconocidos:'},
    'hint':     {'en': 'JSON → promo check  |  +XML → commercial check  |  +Grilla → program check',
                 'es': 'JSON → promos  |  +XML → comerciales  |  +Grilla → programas'},
    'channels': {'en': 'Channels to check:',                        'es': 'Canales a verificar:'},
    'tab_all':  {'en': '📋 All',                                    'es': '📋 Todo'},
    'no_json':  {'en': 'Upload at least one Vipe JSON.',            'es': 'Sube al menos un JSON de Vipe.'},
    'report':   {'en': '📄 Report',                                 'es': '📄 Reporte'},
    'running':  {'en': 'Running checks...',                         'es': 'Verificando...'},
}
def t(k): return L[k][lang]

# ── GLOBAL DISPLAY MAPS (needed in render block outside button) ─────────
CH_DISPLAY = {'catv':'CATV 🌎','tvd':'TVD 📺','latam':'Pasiones Latam 🌹',
              'us':'Pasiones US ⭐','tn':'Fast Todonovelas 📺',
              'hu':'Hola TV US 🤝','hl':'Hola TV Latam 🌍'}

col_title, col_copy = st.columns([3, 1])
with col_title:
    st.title(t('title'))
with col_copy:
    st.markdown(
        f'<div style="text-align:right;padding-top:16px;">'
        f'<span style="font-size:1.05rem;font-weight:700;color:#444;">© 2026 Mauricio Hernandez</span><br>'
        f'<span style="font-size:0.85rem;color:#888;">{APP_VERSION}</span></div>',
        unsafe_allow_html=True)

# ── UPLOAD ────────────────────────────────────────────────────────────────────
if 'uploader_key' not in st.session_state: st.session_state.uploader_key = 0

up_col, clr_col = st.columns([5, 1])
with up_col:
    uploaded = st.file_uploader(t('upload'), accept_multiple_files=True, type=None,
                                key=f'all_files_{st.session_state.uploader_key}')
with clr_col:
    st.write("")
    if st.button("🗑 Clear files", use_container_width=True):
        st.session_state.uploader_key += 1
        st.rerun()
result = detect_files(uploaded) if uploaded else ({}, {}, [], [])
days, grillas, unknown_files, sony_files_raw = result

# ── DETECTION TABLE ───────────────────────────────────────────────────────────
# Count detected vs total
total_uploaded = len(uploaded) if uploaded else 0
total_detected = 0
if uploaded:
    total_detected = (
        sum(len(info['json']) + (1 if info.get('xml') else 0)
            for info in days.values())
        + sum(len(gl) if isinstance(gl, list) else 1 for gl in grillas.values())
        + len(sony_files_raw)
    )

if uploaded:
    st.markdown(f"**{'Detected files' if lang=='en' else 'Archivos detectados'}: {total_detected} / {total_uploaded} uploaded**")
    CH_DISPLAY = {'catv':'CATV 🌎','tvd':'TVD 📺','latam':'Pasiones Latam 🌹',
                  'hu':'Hola TV US 🤝','hl':'Hola TV Latam 🌍',
                  'us':'Pasiones US ⭐','tn':'Fast Todonovelas 📺'}
    rows = []
    for (date_str, channel), info in sorted(days.items()):
        ch = CH_DISPLAY.get(channel, channel.upper())
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(date_str, '%Y-%m-%d')
            short_date = d.strftime('%m/%d')
        except: short_date = date_str
        for jf in info['json']:
            rows.append({'Date': date_str, 'Channel': ch,
                         'Type': f'Playlist {short_date}', 'File': jf.name})
        if info['xml']:
            rows.append({'Date': date_str, 'Channel': ch,
                         'Type': f'Log {short_date}', 'File': info['xml'].name})

    # For grillas, read the week-start Monday date from the file content
    for ch_key, gf_list in grillas.items():
        ch = CH_DISPLAY.get(ch_key, ch_key.upper())
        if not isinstance(gf_list, list):
            gf_list = [gf_list]
        for gf in gf_list:
            grilla_date_str = '(PDF)' if gf.name.lower().endswith('.pdf') else '(Week)'
            if not gf.name.lower().endswith('.pdf'):
                try:
                    from openpyxl import load_workbook
                    import io
                    gf.seek(0)
                    wb = load_workbook(io.BytesIO(gf.read()), read_only=True)
                    gf.seek(0)
                    ws = wb.active
                    rows_g = list(ws.iter_rows(max_row=3, values_only=True))
                    if len(rows_g) > 1:
                        monday_val = rows_g[1][2] if len(rows_g[1]) > 2 else None
                        if monday_val and hasattr(monday_val, 'strftime'):
                            grilla_date_str = monday_val.strftime('%m/%d')
                except: pass
            grilla_label = f'{"Grilla" if lang=="es" else "Grid"} {grilla_date_str}'
            rows.append({'Date': grilla_date_str, 'Channel': ch,
                         'Type': grilla_label, 'File': gf.name})
    for uf in unknown_files:
        rows.append({'Date': '?', 'Channel': '?', 'Type': '?', 'File': uf.name})
    # Sony files — use filename only, never pre-read (preserves file state for pairing)
    for sf in sony_files_raw:
        ch_name = SONY_CHANNEL_MAP.get(sf['code'], sf['code'])
        ch_display = f'{sf["code"]} {SONY_EMOJI.get(sf["code"],"📺")} {ch_name}'
        from checker import _date_from_xml_filename as _dxml
        # Both JSON and XML Sony filenames contain YYYYMMDD
        d = _dxml(sf['file'].name)
        if sf['ftype'] == 'xml':
            type_label = f'Log {d.strftime("%m/%d") if d else "?"}'
        else:
            type_label = f'Playlist {d.strftime("%m/%d") if d else "?"}'
        rows.append({'Date': str(d) if d else '?', 'Channel': ch_display,
                     'Type': type_label, 'File': sf['file'].name})
    if rows:
        import pandas as pd
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     hide_index=True, height=min(500, 35 + len(rows)*35))
    if rows:
        import pandas as pd
        df = pd.DataFrame(rows)
        n_dates    = df["Date"].nunique()
        n_channels = df["Channel"].nunique()
        n_types    = df["Type"].nunique()
        n_files    = len(df)
        st.caption(
            f'Dates ({n_dates})  ·  Channels ({n_channels})  ·  '
            f'Types ({n_types})  ·  Files ({n_files})')
    st.caption(t('hint'))
    st.divider()

# ── CHANNEL SELECTOR ──────────────────────────────────────────────────────────
# Build channel list including Sony codes
available = sorted(set(ch for (_, ch) in days.keys()) | set(grillas.keys()))
sony_codes_present = sorted(set(sf['code'] for sf in sony_files_raw))
all_options = available + sony_codes_present
CH_FORMAT = {'catv':'CATV','tvd':'TVD','latam':'Pasiones Latam',
             'us':'Pasiones US','tn':'Fast Todonovelas',
             'hu':'Hola TV US 🤝','hl':'Hola TV Latam 🌍'}
CH_FORMAT.update({code: f'{code} {SONY_EMOJI.get(code,"📺")} {SONY_CHANNEL_MAP.get(code,code)}' for code in sony_codes_present})

if all_options:
    selected_all = st.multiselect(
        t('channels'), options=all_options, default=all_options,
        format_func=lambda x: CH_FORMAT.get(x, x)
    )
else:
    selected_all = []
selected        = [x for x in selected_all if x in available]
selected_sony   = [x for x in selected_all if x in sony_codes_present]

# ── RUN ───────────────────────────────────────────────────────────────────────
if st.button(t('run'), type='primary', use_container_width=True):
    if not days and not sony_files_raw:
        st.error(t('no_json')); st.stop()

    CH_LABELS = {
        'catv': 'CATV', 'tvd': 'TVD',
        'latam': 'Pasiones Latam', 'us': 'Pasiones US',
        'tn': 'Fast Todonovelas',
        'hu': 'Hola TV US 🤝', 'hl': 'Hola TV Latam 🌍',
    }

    def process_one(channel, json_file, xml_file, grilla_file):
        is_tn  = (channel == 'tn')
        lines  = []
        file_info  = {
            'json':   json_file.name  if json_file   else None,
            'xml':    xml_file.name   if xml_file    else None,
            'grilla': grilla_file.name if grilla_file else None,
        }
        try:
            json_file.seek(0)
            data     = json.load(json_file)
            playlist = parse_json_playlist(data)
        except Exception as e:
            return [f'ERROR parsing JSON: {e}'], []
        xml_rows = []
        if xml_file:
            try:
                xml_file.seek(0)
                xml_rows = parse_xml_log_tn(xml_file) if is_tn else parse_xml_log(xml_file)
            except Exception as e:
                lines.append(f'  WARNING: XML error: {e}')
        grilla_ids = []
        if grilla_file and playlist['date']:
            try:
                grilla_file.seek(0)
                grilla_ids = parse_grilla(grilla_file, playlist['date'], channel)
            except Exception as e:
                lines.append(f'  WARNING: Grilla error: {e}')
        if not xml_rows and not grilla_ids:
            pi = check_promo_repeats(playlist, lang=lang)
            ch_label = CH_LABELS.get(channel, channel.upper())
            lines += ['═'*60,
                      f'CHANNEL: {ch_label} | DATE: {playlist["date"]} | {"PROMO CHECK" if lang=="en" else "VERIFICACIÓN DE PROMOS"}',
                      '═'*60]
            lines += pi if pi else ['  ✓ No repeated promos']
            lines.append('')
            return lines, []
        ch_label = CH_LABELS.get(channel, channel.upper())
        try:
            report_text, manual_warns = generate_report(
                ch_label, playlist, xml_rows, grilla_ids, lang,
                is_tn=is_tn, file_info=file_info)
            lines.append(report_text)
        except Exception as e:
            lines.append(f'  ERROR generating report: {e}')
            manual_warns = []
        return lines, manual_warns

    # Build per-day reports
    sorted_dates = sorted(set(d for (d, _) in days.keys()))
    day_reports  = {}  # date_str -> list of lines
    header_lines = [
        'BROADCAST PLAYLIST CHECK REPORT' if lang=='en' else 'REPORTE DE VERIFICACIÓN DE PLAYLIST',
        f'Generated / Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        '═'*60, ''
    ]

    all_manual_warns = []  # list of (channel_label, date_str, count)
    CH_DISPLAY = {'catv':'CATV 🌎','tvd':'TVD 📺','latam':'Pasiones Latam 🌹',
                  'hu':'Hola TV US 🤝','hl':'Hola TV Latam 🌍','us':'Pasiones US ⭐','tn':'Fast Todonovelas 📺'}
    with st.spinner(t('running')):
        for date_str in sorted_dates:
            d_lines = [f'{"DATE" if lang=="en" else "FECHA"}: {date_str}', '─'*60]
            ch_reports = {}  # channel -> lines
            for channel in ['catv', 'tvd', 'latam', 'us', 'tn', 'hu', 'hl']:
                if channel not in selected: continue
                key = (date_str, channel)
                if key not in days: continue
                info     = days[key]
                xml_file = info.get('xml')
                # grillas stores a list — pick the one whose week matches this date
                gf_list  = grillas.get(channel, [])
                if not isinstance(gf_list, list): gf_list = [gf_list] if gf_list else []
                grilla_warn = None
                try:
                    from datetime import datetime as _dt2
                    from checker import pick_grilla_for_date
                    td = _dt2.strptime(date_str, '%Y-%m-%d').date()
                    grilla_f, grilla_warn = pick_grilla_for_date(gf_list, td, channel)
                except:
                    grilla_f = gf_list[0] if gf_list else None
                # Sort: partial (no marker) before full (has marker)
                def _is_full(jf):
                    try:
                        jf.seek(0)
                        d = json.load(jf)
                        jf.seek(0)
                        return any(a.get('type')=='marker'
                                   for ev in d.get('events',[])[:3]
                                   for a in ev.get('assets',[]))
                    except: jf.seek(0); return False
                jsons = sorted(info['json'], key=lambda f: (1 if _is_full(f) else 0, f.name))
                ch_lines = []
                if grilla_warn:
                    ch_lines.append(f'  ⚠  {grilla_warn}')
                for jf in jsons:
                    ch_lines.append(f'JSON: {jf.name}')
                    # ── HolaTV channels ──────────────────────────────────
                    if channel in ('hu', 'hl'):
                        try:
                            jf.seek(0)
                            _pl  = parse_json_playlist(json.load(jf))
                            jf.seek(0)
                            _lf  = info.get('xml') or info.get('log')
                            _lr, _dx, _cx = (load_holatv_log(_lf, info['date'])
                                             if _lf else ([], 0, 0))
                            _gfl = grillas.get(channel, [])
                            if not isinstance(_gfl, list): _gfl = [_gfl] if _gfl else []
                            _gf_raw = pick_grilla_for_date(_gfl, info['date'], channel)
                            _gf  = _gf_raw[0] if isinstance(_gf_raw, tuple) else _gf_raw
                            _ge  = (parse_grilla_holatv_v2(_gf, info['date'])
                                    if _gf and _gf.name.lower().endswith('.pdf') else [])
                            _cs  = (_pl['programs'][0]['start']
                                    if _pl['type']=='current' and _pl['programs'] else None)
                            _fi  = {'json': jf.name,
                                    'log':  _lf.name if _lf else None,
                                    'grilla': _gf.name if _gf else None}
                            ch_lines.append(generate_report_holatv_v2(
                                CH_LABELS.get(channel, channel.upper()),
                                _lr, _dx, _cx, _ge, _pl, lang,
                                file_info=_fi, current_start_utc=_cs))
                        except Exception as _he:
                            ch_lines.append(f'  ERROR (HolaTV): {_he}')
                        continue
                    # ─────────────────────────────────────────────────────
                    if xml_file:  xml_file.seek(0)
                    if grilla_f:  grilla_f.seek(0)
                    result = process_one(channel, jf, xml_file, grilla_f)
                    plines, warns = result if isinstance(result, tuple) else (result, [])
                    ch_lines += plines
                    if warns:
                        all_manual_warns.append(
                            (CH_DISPLAY.get(channel, channel.upper()), date_str, len(warns))
                        )
                ch_reports[channel] = ch_lines
                d_lines += ch_lines
            d_lines.append('')
            day_reports[date_str] = {'all': d_lines, 'channels': ch_reports}

    # ── SONY / AXN PROCESSING (single pass) ──────────────────────────────────
    # Group results by (date, code) to integrate into date tabs like other channels
    sony_date_ch = {}   # {date_str: {code: lines}}  for tab integration
    sony_by_code = {}   # {code: lines}               for dedicated Sony tabs
    if sony_files_raw and selected_sony:
        with st.spinner(t('running')):
            sony_pairings = pair_sony_files(sony_files_raw, lang)
        sep60 = '═' * 60
        for pair in sony_pairings:
            if pair['code'] not in selected_sony:
                continue
            code     = pair['code']
            date_str = str(pair['date']) if pair['date'] else '?'
            markers_in_json = parse_sony_json_markers(pair['json_data']) if pair['json_data'] else []
            if pair['json_data'] is None:   pl_type = SL('pl_none', lang)
            elif markers_in_json:           pl_type = SL('pl_full', lang)
            else:                           pl_type = SL('pl_partial', lang)

            pairing_lines  = [sep60,
                              f'CHANNEL: {code} — {pair["channel_name"]}',
                              f'DATE: {date_str}',
                              f'PLAYLIST TYPE: {pl_type}',
                              f'JSON: {pair["json_file"].name if pair["json_file"] else "— not provided —"}',
                              f'LOG:  {pair["xml_filename"] or "— not provided —"}',
                              sep60]

            if pair['json_data'] is None and pair['xml_file']:
                pairing_lines.append(SL('no_json', lang))
            elif pair['json_data'] is not None and pair['xml_file'] is None:
                pairing_lines.append(SL('no_log', lang))
                try:
                    r_lines, _ = check_sony(pair['json_data'], [], None, lang)
                    pairing_lines += r_lines
                except: pass
            elif pair['json_data'] is not None and pair['xml_file'] is not None:
                try:
                    pair['xml_file'].seek(0)
                    xml_rows_sony = parse_sony_xml_log(pair['xml_file'])
                    r_lines, _ = check_sony(pair['json_data'], xml_rows_sony,
                                            pair['xml_filename'], lang)
                    pairing_lines += r_lines
                except Exception as e:
                    pairing_lines.append(f'  ERROR: {e}')
            pairing_lines.append('')

            # Store under date bucket for tab integration
            sony_date_ch.setdefault(date_str, {})
            sony_date_ch[date_str].setdefault(code, [])
            sony_date_ch[date_str][code] += pairing_lines
            # Also store per-code for dedicated tabs
            sony_by_code.setdefault(code, [])
            sony_by_code[code] += pairing_lines

    # ── BUILD FULL REPORT TEXT ────────────────────────────────────────────────
    st.session_state['_report_ready'] = True
    all_lines = header_lines[:]
    for d in sorted_dates:
        all_lines += day_reports.get(d, {}).get('all', [])
    # Add Sony grouped by date
    for date_str, codes in sorted(sony_date_ch.items()):
        all_lines += [f'{"DATE" if lang=="en" else "FECHA"}: {date_str} (Sony/AXN)', '─'*60]
        for code, lines in sorted(codes.items()):
            all_lines += lines
    full_text = '\n'.join(all_lines)

    # Store all report data in session state — survives widget interactions
    st.session_state['report_full_text']    = full_text
    st.session_state['report_day_reports']   = day_reports
    st.session_state['report_sorted_dates']  = sorted_dates
    st.session_state['report_header_lines']  = header_lines
    st.session_state['report_sony_date_ch']  = sony_date_ch
    st.session_state['report_sony_by_code']  = sony_by_code
    st.session_state['report_all_warns']     = all_manual_warns
    st.session_state['report_lang']          = lang


# ── REPORT RENDER (outside button block — survives reruns) ──────────────────
if st.session_state.get('_report_ready'):
    full_text    = st.session_state.get('report_full_text', '')
    day_reports  = st.session_state.get('report_day_reports', {})
    sorted_dates = st.session_state.get('report_sorted_dates', [])
    header_lines = st.session_state.get('report_header_lines', [])
    sony_date_ch = st.session_state.get('report_sony_date_ch', {})
    sony_by_code = st.session_state.get('report_sony_by_code', {})
    lang         = st.session_state.get('report_lang', 'en')

    st.subheader('📄 Report' if lang=='en' else '📄 Reporte')

    # ── Manual review warnings
    all_manual_warns = st.session_state.get('report_all_warns', [])
    if all_manual_warns:
        warn_header = '⚠ MANUAL REVIEW NEEDED:' if lang == 'en' else '⚠ REVISIÓN MANUAL REQUERIDA:'
        warn_lines = [warn_header]
        for ch_lbl, d_str, cnt in all_manual_warns:
            warn_lines.append('')
            warn_lines.append(f'  {ch_lbl}  —  {d_str}')
            warn_lines.append(f'    {cnt} commercial block{"s" if cnt>1 else ""} need manual review')
        st.error('\n'.join(warn_lines))

    # ── Empty report detection
    empty_reports = []
    for date_str in sorted_dates:
        ch_rpts = day_reports.get(date_str, {}).get('channels', {})
        for ch, lines in ch_rpts.items():
            text = '\n'.join(lines)
            has_content = any(s in text for s in ['SUMMARY', 'PROGRAM CHECK', 'COMMERCIAL CHECK'])
            if not has_content:
                empty_reports.append((CH_DISPLAY.get(ch, ch), date_str))
    if empty_reports:
        empty_lbl = 'Empty Reports:' if lang == 'en' else 'Reportes Vacíos:'
        e_lines = [f'ℹ  {empty_lbl}']
        for ch_lbl, d_str in empty_reports:
            e_lines.append(f'  {ch_lbl} — {d_str}')
        st.warning('\n'.join(e_lines))

    # ── Not Ingested only (expander with channel filter + download)
    with st.expander('🔍 Not Ingested only' if lang=='en' else '🔍 Solo No Ingestados'):
        # Build per-channel NI text from full report
        _ni_by_ch = {}
        _cur_ch = None
        _ft = full_text
        for _line in _ft.splitlines():
            if _line.startswith('CHANNEL:') or _line.startswith('CANAL:'):
                _cur_ch = _line
            if 'NOT INGESTED' in _line or 'NO INGESTADO' in _line:
                if _cur_ch not in _ni_by_ch: _ni_by_ch[_cur_ch] = []
                _ni_by_ch[_cur_ch].append(_line)
        if _ni_by_ch:
            _ni_opts = list(_ni_by_ch.keys())
            _ni_sel  = st.multiselect(
                'Filter by channel' if lang=='en' else 'Filtrar por canal',
                options=_ni_opts, default=_ni_opts,
                format_func=lambda x: x.replace('CHANNEL: ','').replace('CANAL: ',''),
                key='ni_ch_filter')
            _ni_lines = []
            for _ch in _ni_sel:
                _ni_lines.append(_ch)
                _ni_lines += _ni_by_ch[_ch]
                _ni_lines.append('')
            _ni_text = '\n'.join(_ni_lines)
            st.text(_ni_text)
            st.download_button(
                '⬇ Download Not Ingested (.txt)' if lang=='en' else '⬇ Descargar No Ingestados (.txt)',
                _ni_text,
                file_name=f'not_ingested_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt',
                mime='text/plain', use_container_width=True, key='dl_ni')
        else:
            st.write('None found.' if lang=='en' else 'Ninguno encontrado.')

    # Build tab structure: All + per-date (regular + Sony mixed) + dedicated Sony tabs
    CH_DISPLAY2 = {**CH_DISPLAY,
                   **{code: f'{code} {SONY_EMOJI.get(code,"📺")} {SONY_CHANNEL_MAP.get(code,code)}'
                      for code in st.session_state.get('report_sony_by_code',{})}}

    # Merge Sony channels into date reports for unified date tabs
    all_tab_dates = sorted(set(list(sorted_dates) + list(sony_date_ch.keys())))

    def _render_day_tab(date_str, key_prefix):
        day_data   = st.session_state.get('report_day_reports',"").get(date_str, {})
        ch_reports = dict(day_data.get('channels', {}))
        # Add Sony channels for this date
        sony_for_date = st.session_state.get('report_sony_date_ch',{}).get(date_str, {})
        for code, lines in sorted(sony_for_date.items()):
            ch_reports[code] = lines
        day_text = '\n'.join(header_lines + day_data.get('all', []) +
                              [l for lines in sony_for_date.values() for l in lines])
        if len(ch_reports) > 1:
            ch_tab_labels = [CH_DISPLAY2.get(ch, ch) for ch in ch_reports]
            ch_tabs = st.tabs(ch_tab_labels)
            for j, (ch, ch_lines) in enumerate(ch_reports.items()):
                with ch_tabs[j]:
                    ch_text = '\n'.join(header_lines + ch_lines)
                    st.text(ch_text)
                    st.download_button(t('dl'), ch_text,
                                       file_name=f'report_{date_str}_{ch}_{datetime.now().strftime("%H%M%S")}.txt',
                                       mime='text/plain', use_container_width=True,
                                       key=f'dl_{key_prefix}_{date_str}_{ch}')
        else:
            st.text(day_text)
        dl_day_lbl = f'⬇ {"Current" if lang=="en" else "Actual"} {date_str} (.txt)'
        st.download_button(dl_day_lbl, day_text,
                           file_name=f'report_{date_str}_{datetime.now().strftime("%H%M%S")}.txt',
                           mime='text/plain', use_container_width=True,
                           key=f'dl_{key_prefix}_{date_str}_all')

    if all_tab_dates:
        tab_labels = [t('tab_all')] + [f'📅 {d}' for d in all_tab_dates]

        tabs = st.tabs(tab_labels)

        # All tab
        with tabs[0]:
            st.text(full_text)
            dl_lbl = '⬇ Full Report (.txt)' if lang=='en' else '⬇ Reporte Completo (.txt)'
            st.download_button(dl_lbl, st.session_state.get('report_full_text',""),
                               file_name=f'report_all_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt',
                               mime='text/plain', use_container_width=True, key='dl_all')

        # Date tabs (regular channels + Sony merged by date)
        for i, date_str in enumerate(all_tab_dates):
            with tabs[i+1]:
                _render_day_tab(date_str, 'day')
    else:
        st.text(full_text)
        dl_lbl3 = '⬇ Full Report (.txt)' if lang=='en' else '⬇ Reporte Completo (.txt)'
        st.download_button(dl_lbl3, st.session_state.get('report_full_text',""),
                           file_name=f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt',
                           mime='text/plain', use_container_width=True, key='dl_single_all')
